#!/usr/bin/env python3
"""Inspect saved SQL result files without invoking review/LLM flows."""

from __future__ import annotations

import csv
import hashlib
import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from sql_time_contract import (
    actual_range_role,
    is_temporal_output_name,
    output_field_key,
    parse_datetime,
    project_time_integrity_policy,
    project_time_policy,
    requested_time_window,
)


RESULT_EXTENSIONS = {".csv", ".tsv", ".xlsx"}
RESULT_TIME_COVERAGE_VERSION = "result_time_coverage_v1"
DATE_LIKE_VALUE_RE = re.compile(
    r"^\d{4}(?:[-/]\d{1,2}[-/]\d{1,2}|\d{4})"
    r"(?:[ T]\d{1,2}:\d{1,2}(?::\d{1,2}(?:\.\d+)?)?)?"
    r"(?:Z|[+-]\d{2}:?\d{2})?$"
)
SUMMARY_LABELS = {"合计", "总计", "整体", "total", "overall", "all"}
RATIO_FIELD_RE = re.compile(
    r"(占比|比例|比率|转化率|留存率|率|percent(?:age)?|ratio|rate|share|(?:^|[_\s])pct(?:$|[_\s]))",
    re.I,
)


def _compact(value: Any, limit: int = 240) -> str:
    text = " ".join(str(value or "").split())
    if len(text) <= limit:
        return text
    return text[: limit - 1].rstrip() + "..."


def _candidate_time_fields(columns: list[str], preferred_fields: list[str]) -> list[str]:
    preferred = {output_field_key(item) for item in preferred_fields if str(item or "").strip()}
    rows: list[str] = []
    for column in columns:
        if output_field_key(column) in preferred or is_temporal_output_name(column):
            rows.append(column)
    return rows


def _observed_datetime(value: Any) -> tuple[datetime | None, str, bool]:
    if value is None:
        return None, "missing", False
    if isinstance(value, datetime):
        parsed = value.replace(tzinfo=None)
        has_time_precision = True
    elif isinstance(value, date):
        parsed = datetime(value.year, value.month, value.day)
        has_time_precision = False
    else:
        text = str(value).strip()
        if not text:
            return None, "missing", False
        if text.casefold() in SUMMARY_LABELS:
            return None, "summary_label", False
        normalized = text.replace("T", " ")
        if not DATE_LIKE_VALUE_RE.fullmatch(normalized):
            return None, "invalid_format", False
        parsed = parse_datetime(normalized)
        if not parsed:
            return None, "invalid_calendar", False
        has_time_precision = bool(re.search(r"\s\d{1,2}:\d{1,2}", normalized))
    if not 1970 <= parsed.year <= 2100:
        return None, "non_gregorian_or_outlier_year", has_time_precision
    return parsed, "valid", has_time_precision


class _TimeCoverageCollector:
    def __init__(
        self,
        columns: list[str],
        *,
        preferred_fields: list[str],
        requested_window: dict[str, Any],
    ) -> None:
        self.columns = _candidate_time_fields(columns, preferred_fields)
        self.requested_window = requested_window
        self.rows: dict[str, dict[str, Any]] = {
            column: {
                "field": column,
                "valid_count": 0,
                "parsed_count": 0,
                "datetime_value_count": 0,
                "invalid_count": 0,
                "missing_count": 0,
                "summary_label_count": 0,
                "outside_requested_count": 0,
                "actual_start": "",
                "actual_end": "",
                "invalid_examples": [],
            }
            for column in self.columns
        }
        self._minimums: dict[str, datetime] = {}
        self._maximums: dict[str, datetime] = {}
        self._requested_start = parse_datetime(requested_window.get("start_date"))
        self._requested_end = parse_datetime(requested_window.get("end_date"))
        self._comparison_precision = str(
            requested_window.get("comparison_precision") or "date"
        )
        self._comparison_start = parse_datetime(
            requested_window.get("comparison_start")
        )
        self._comparison_end = parse_datetime(
            requested_window.get("comparison_end")
        )
        self._comparison_start_operator = str(
            requested_window.get("comparison_start_operator") or ">="
        )
        self._comparison_end_operator = str(
            requested_window.get("comparison_end_operator") or "<="
        )

    def _outside_requested_window(self, parsed: datetime) -> bool:
        if (
            self._comparison_precision == "datetime"
            and self._comparison_start
            and self._comparison_end
        ):
            before_start = (
                parsed <= self._comparison_start
                if self._comparison_start_operator == ">"
                else parsed < self._comparison_start
            )
            after_end = (
                parsed >= self._comparison_end
                if self._comparison_end_operator == "<"
                else parsed > self._comparison_end
            )
            return before_start or after_end
        if self._requested_start and parsed.date() < self._requested_start.date():
            return True
        return bool(
            self._requested_end and parsed.date() > self._requested_end.date()
        )

    def add(self, row: dict[str, Any]) -> None:
        for column in self.columns:
            value = row.get(column)
            parsed, status, has_time_precision = _observed_datetime(value)
            item = self.rows[column]
            if status == "missing":
                item["missing_count"] += 1
                continue
            if status == "summary_label":
                item["summary_label_count"] += 1
                continue
            if not parsed:
                item["invalid_count"] += 1
                if len(item["invalid_examples"]) < 3:
                    item["invalid_examples"].append(
                        {"value": _compact(value, 80), "reason": status}
                    )
                continue
            item["parsed_count"] += 1
            if has_time_precision:
                item["datetime_value_count"] += 1
            outside_requested = self._outside_requested_window(parsed)
            if outside_requested:
                item["outside_requested_count"] += 1
                if len(item["invalid_examples"]) < 3:
                    item["invalid_examples"].append(
                        {"value": _compact(value, 80), "reason": "outside_requested_window"}
                    )
                continue
            item["valid_count"] += 1
            self._minimums[column] = min(self._minimums.get(column, parsed), parsed)
            self._maximums[column] = max(self._maximums.get(column, parsed), parsed)

    def result(self, row_count: int, preferred_fields: list[str]) -> dict[str, Any]:
        preferred = {output_field_key(item) for item in preferred_fields if str(item or "").strip()}
        fields: list[dict[str, Any]] = []
        for column in self.columns:
            item = dict(self.rows[column])
            minimum = self._minimums.get(column)
            maximum = self._maximums.get(column)
            precision = "datetime" if item["datetime_value_count"] else "date"
            item["precision"] = precision
            item["actual_start"] = (
                minimum.isoformat(sep=" ")
                if minimum and precision == "datetime"
                else minimum.date().isoformat()
                if minimum
                else ""
            )
            item["actual_end"] = (
                maximum.isoformat(sep=" ")
                if maximum and precision == "datetime"
                else maximum.date().isoformat()
                if maximum
                else ""
            )
            fields.append(item)
        observed = [item for item in fields if item["valid_count"] > 0]
        observed.sort(
            key=lambda item: (
                item["valid_count"],
                item["precision"] == "datetime",
                output_field_key(item["field"]) in preferred,
                -item["invalid_count"],
            ),
            reverse=True,
        )
        range_starts = [item for item in observed if actual_range_role(item["field"]) == "start"]
        range_ends = [item for item in observed if actual_range_role(item["field"]) == "end"]
        explicit_start = range_starts[0] if range_starts else None
        explicit_end = range_ends[0] if range_ends else None
        explicit_range = bool(explicit_start and explicit_end)
        primary = observed[0] if observed else None
        selected_fields = (
            [explicit_start, explicit_end]
            if explicit_range
            else ([primary] if primary else [])
        )
        selected_anomaly_count = sum(
            int(item.get("invalid_count") or 0)
            + int(item.get("outside_requested_count") or 0)
            for item in selected_fields
            if item
        )
        if row_count == 0:
            status = "empty"
        elif primary is None:
            status = "not_observable"
        elif selected_anomaly_count:
            status = "observed_with_anomalies"
        else:
            status = "observed"
        required = bool(
            self.requested_window.get("today_included") is True
            or self.requested_window.get("dynamic")
        )
        if explicit_range:
            actual_start = str(explicit_start.get("actual_start") or "")
            actual_end = str(explicit_end.get("actual_end") or "")
            basis = "explicit_range_fields"
            primary_field = ""
            range_fields = {
                "start": str(explicit_start.get("field") or ""),
                "end": str(explicit_end.get("field") or ""),
            }
            precision = (
                "datetime"
                if explicit_start.get("precision") == "datetime"
                and explicit_end.get("precision") == "datetime"
                else "date"
            )
            if precision == "date":
                parsed_start = parse_datetime(actual_start)
                parsed_end = parse_datetime(actual_end)
                actual_start = parsed_start.date().isoformat() if parsed_start else actual_start
                actual_end = parsed_end.date().isoformat() if parsed_end else actual_end
        else:
            actual_start = primary["actual_start"] if primary else ""
            actual_end = primary["actual_end"] if primary else ""
            basis = "result_output_field" if primary else "not_observable_from_result"
            primary_field = primary["field"] if primary else ""
            range_fields = {"start": "", "end": ""}
            precision = str(primary.get("precision") or "none") if primary else "none"
        if actual_start and actual_end:
            start_value = parse_datetime(actual_start)
            end_value = parse_datetime(actual_end)
            if start_value and end_value and end_value < start_value:
                status = "observed_with_anomalies"
                selected_anomaly_count += 1
        if not required:
            requirement_status = "not_required"
        elif status == "empty":
            requirement_status = "met_empty"
        elif status == "observed":
            requirement_status = "met"
        elif primary is None:
            requirement_status = "not_observable"
        else:
            requirement_status = "anomalous"
        return {
            "schema_version": RESULT_TIME_COVERAGE_VERSION,
            "status": status,
            "basis": basis,
            "required": required,
            "requirement_status": requirement_status,
            "requested_window": dict(self.requested_window),
            "primary_field": primary_field,
            "range_fields": range_fields,
            "precision": precision,
            "actual_start": actual_start,
            "actual_end": actual_end,
            "excluded_anomaly_count": selected_anomaly_count,
            "fields": fields,
            "field_count": len(fields),
        }


def time_coverage_problem_messages(coverage: dict[str, Any] | None) -> list[str]:
    coverage = coverage if isinstance(coverage, dict) else {}
    if not coverage.get("required"):
        return []
    status = str(coverage.get("requirement_status") or "not_observable")
    if status in {"met", "met_empty"}:
        return []
    if status == "anomalous":
        return [
            "查询范围包含或可能包含今日，但结果中的实际数据时间范围仍包含无效历法、请求区间外值或倒置边界。"
        ]
    return [
        "查询范围包含或可能包含今日，但结果无法观测实际数据时间范围；请从已过滤基础层输出日期/时间字段，"
        "或输出 `实际数据开始时间` 与 `实际数据结束时间`。"
    ]


def unobservable_time_coverage(
    *,
    sql: str,
    project_config: dict[str, Any] | None,
    as_of_date: str | date | None = None,
    basis: str = "unsupported_result_format",
) -> dict[str, Any]:
    window = requested_time_window(sql, project_config or {}, as_of_date=as_of_date)
    required = bool(window.get("today_included") is True or window.get("dynamic"))
    return {
        "schema_version": RESULT_TIME_COVERAGE_VERSION,
        "status": "not_observable",
        "basis": basis,
        "required": required,
        "requirement_status": "not_observable" if required else "not_required",
        "requested_window": window,
        "primary_field": "",
        "range_fields": {"start": "", "end": ""},
        "precision": "none",
        "actual_start": "",
        "actual_end": "",
        "excluded_anomaly_count": 0,
        "fields": [],
        "field_count": 0,
    }


def _read_csv(
    path: Path,
    sample_limit: int,
    *,
    preferred_time_fields: list[str],
    time_window: dict[str, Any],
) -> tuple[list[str], list[dict[str, Any]], int, dict[str, Any]]:
    encodings = ["utf-8-sig", "utf-8", "gb18030"]
    last_error: Exception | None = None
    for encoding in encodings:
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                reader = csv.DictReader(handle, delimiter="\t" if path.suffix.lower() == ".tsv" else ",")
                columns = [str(item or "").strip() for item in (reader.fieldnames or [])]
                coverage = _TimeCoverageCollector(
                    columns,
                    preferred_fields=preferred_time_fields,
                    requested_window=time_window,
                )
                rows: list[dict[str, Any]] = []
                count = 0
                for row in reader:
                    count += 1
                    coverage.add(row)
                    if len(rows) < sample_limit:
                        rows.append({key: _compact(value) for key, value in row.items()})
                return columns, rows, count, coverage.result(count, preferred_time_fields)
        except UnicodeDecodeError as exc:
            last_error = exc
            continue
    if last_error:
        raise last_error
    return [], [], 0, _TimeCoverageCollector(
        [], preferred_fields=preferred_time_fields, requested_window=time_window
    ).result(0, preferred_time_fields)


def _read_xlsx(
    path: Path,
    sample_limit: int,
    *,
    preferred_time_fields: list[str],
    time_window: dict[str, Any],
) -> tuple[list[str], list[dict[str, Any]], int, dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - runtime dependency guard
        raise ValueError("XLSX result inspection requires openpyxl.") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        columns: list[str] = []
        rows: list[dict[str, Any]] = []
        count = 0
        coverage: _TimeCoverageCollector | None = None
        for raw_values in worksheet.iter_rows(values_only=True):
            raw = list(raw_values)
            if not any(str(value or "").strip() for value in raw):
                continue
            if not columns:
                columns = [
                    str(value or "").strip() or f"column_{index + 1}"
                    for index, value in enumerate(raw)
                ]
                coverage = _TimeCoverageCollector(
                    columns,
                    preferred_fields=preferred_time_fields,
                    requested_window=time_window,
                )
                continue
            count += 1
            raw_row = {
                columns[index]: raw[index] if index < len(raw) else None
                for index in range(len(columns))
            }
            if coverage:
                coverage.add(raw_row)
            if len(rows) < sample_limit:
                rows.append(
                    {
                        columns[index]: _compact(raw[index] if index < len(raw) else "")
                        for index in range(len(columns))
                    }
                )
        coverage = coverage or _TimeCoverageCollector(
            columns,
            preferred_fields=preferred_time_fields,
            requested_window=time_window,
        )
        return columns, rows, count, coverage.result(count, preferred_time_fields)
    finally:
        workbook.close()


def _to_number(value: Any) -> float | None:
    text = str(value or "").strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def ratio_field_rules(columns: list[str], rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rules: list[dict[str, Any]] = []
    for column in columns:
        if not RATIO_FIELD_RE.search(column):
            continue
        numbers = [_to_number(row.get(column)) for row in rows]
        numbers = [value for value in numbers if value is not None]
        scale = "ratio_0_to_1"
        raw_value: float | str | None = numbers[0] if numbers else None
        if numbers and max(abs(value) for value in numbers) > 1:
            scale = "percent_0_to_100"
        if isinstance(raw_value, float):
            display_number = raw_value * 100 if scale == "ratio_0_to_1" else raw_value
            display_value = f"{display_number:.2f}%"
        else:
            display_value = "示例值缺失"
        rules.append(
            {
                "output_field": column,
                "semantic_type": "ratio",
                "source_value_scale": scale,
                "display_format": "percent",
                "decimal_places": 2,
                "display_suffix": "%",
                "preserve_raw_value": True,
                "sample_check": {"raw_value": raw_value, "display_value": display_value},
                "review_note": "快线根据结果字段名和值域自动识别；SQL 原始值保持不变，DA 展示为百分比两位小数。",
            }
        )
    return rules


def inspect_result_file(
    path: Path,
    sample_limit: int = 8,
    *,
    sql: str = "",
    project_config: dict[str, Any] | None = None,
    as_of_date: str | date | None = None,
    preferred_time_fields: list[str] | None = None,
) -> dict[str, Any]:
    path = path.resolve()
    if not path.exists():
        raise FileNotFoundError(path)
    suffix = path.suffix.lower()
    if suffix not in RESULT_EXTENSIONS:
        raise ValueError(f"Unsupported result file type: {suffix}")
    config = project_config or {}
    partition = project_time_policy(config)
    integrity = project_time_integrity_policy(config)
    preferred_fields = [
        *(preferred_time_fields or []),
        str(integrity.get("date_field") or ""),
        str(integrity.get("time_field") or ""),
        str(partition.get("partition_field") or ""),
        str(partition.get("business_time_field") or ""),
    ]
    time_window = (
        requested_time_window(sql, config, as_of_date=as_of_date)
        if sql
        else {
            "start_date": "",
            "end_date": "",
            "comparison_precision": "unknown",
            "comparison_start": "",
            "comparison_end": "",
            "comparison_start_operator": "",
            "comparison_end_operator": "",
            "comparison_field": "",
            "basis": "not_linked_to_sql",
            "dynamic": False,
            "today_included": False,
            "as_of_date": str(as_of_date or ""),
            "timezone_offset": str(integrity.get("timezone_offset") or "+08:00"),
        }
    )
    if suffix in {".csv", ".tsv"}:
        columns, rows, row_count, time_coverage = _read_csv(
            path,
            sample_limit,
            preferred_time_fields=preferred_fields,
            time_window=time_window,
        )
    else:
        columns, rows, row_count, time_coverage = _read_xlsx(
            path,
            sample_limit,
            preferred_time_fields=preferred_fields,
            time_window=time_window,
        )
    fingerprint_payload = {"columns": columns, "sample": rows[:3], "row_count": row_count}
    return {
        "path": str(path),
        "file_name": path.name,
        "file_type": suffix.lstrip("."),
        "row_count": row_count,
        "columns": columns,
        "sample_rows": rows,
        "ratio_field_rules": ratio_field_rules(columns, rows),
        "time_coverage": time_coverage,
        "schema_fingerprint": hashlib.sha256(
            json.dumps(fingerprint_payload, ensure_ascii=False, sort_keys=True).encode("utf-8")
        ).hexdigest()[:16],
    }
