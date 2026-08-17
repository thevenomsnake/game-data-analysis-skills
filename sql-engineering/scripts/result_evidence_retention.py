#!/usr/bin/env python3
"""Retain compact SQL result evidence while preserving reusable outputs in full."""

from __future__ import annotations

import csv
import hashlib
import io
import mimetypes
import shutil
from collections import deque
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable


RESULT_EVIDENCE_MAX_BYTES = 10 * 1024 * 1024
RESULT_EVIDENCE_SLICE_MAX_BYTES = 8 * 1024 * 1024
RESULT_EVIDENCE_HEAD_ROWS = 500
RESULT_EVIDENCE_TAIL_ROWS = 100
RESULT_EVIDENCE_RETENTION_VERSION = "result_evidence_retention_v1"
REUSABLE_OUTPUT_KINDS = {
    "analysis_workbook",
    "comparison_workbook",
    "visualization",
    "export",
    "other",
}
SLICEABLE_RESULT_EXTENSIONS = {".csv", ".tsv", ".xlsx"}


@dataclass(frozen=True)
class RetainedResult:
    payload: bytes
    suffix: str
    media_type: str
    stored_sha256: str
    retention: dict[str, Any]


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _detect_text_encoding(path: Path) -> str:
    with path.open("rb") as handle:
        sample = handle.read(65536)
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            sample.decode(encoding)
            return encoding
        except UnicodeDecodeError:
            continue
    raise ValueError(f"Cannot decode result evidence as UTF-8/GB18030: {path}")


def _detect_delimiter(path: Path, encoding: str) -> str:
    if path.suffix.lower() == ".tsv":
        return "\t"
    with path.open("r", encoding=encoding, newline="") as handle:
        sample = handle.read(32768)
    try:
        return csv.Sniffer().sniff(sample, delimiters=",\t;|").delimiter
    except csv.Error:
        return ","


def _normalized_cell(value: Any) -> Any:
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except (TypeError, ValueError):
            pass
    return value


def _collect_head_tail(rows: Iterable[Iterable[Any]]) -> tuple[list[Any], list[list[Any]], int]:
    header: list[Any] = []
    head: list[list[Any]] = []
    tail: deque[list[Any]] = deque(maxlen=RESULT_EVIDENCE_TAIL_ROWS)
    row_count = 0
    for raw_row in rows:
        row = [_normalized_cell(value) for value in raw_row]
        if not any(str(value).strip() for value in row):
            continue
        if not header:
            header = row
            continue
        row_count += 1
        if len(head) < RESULT_EVIDENCE_HEAD_ROWS:
            head.append(row)
        else:
            tail.append(row)
    return header, head + list(tail), row_count


def _csv_rows(path: Path) -> tuple[list[Any], list[list[Any]], int, str]:
    encoding = _detect_text_encoding(path)
    delimiter = _detect_delimiter(path, encoding)
    with path.open("r", encoding=encoding, newline="") as handle:
        header, rows, row_count = _collect_head_tail(csv.reader(handle, delimiter=delimiter))
    return header, rows, row_count, delimiter


def _xlsx_rows(path: Path) -> tuple[list[Any], list[list[Any]], int, str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - runtime dependency guard
        raise ValueError("Large XLSX result slicing requires openpyxl.") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        header, rows, row_count = _collect_head_tail(
            worksheet.iter_rows(values_only=True)
        )
        return header, rows, row_count, worksheet.title
    finally:
        workbook.close()


def _row_bytes(row: list[Any], delimiter: str) -> bytes:
    buffer = io.StringIO(newline="")
    csv.writer(buffer, delimiter=delimiter, lineterminator="\n").writerow(row)
    return buffer.getvalue().encode("utf-8")


def _render_slice(header: list[Any], rows: list[list[Any]], delimiter: str) -> tuple[bytes, int]:
    output = bytearray(b"\xef\xbb\xbf")
    header_bytes = _row_bytes(header, delimiter)
    if len(output) + len(header_bytes) > RESULT_EVIDENCE_SLICE_MAX_BYTES:
        raise ValueError("Result header alone exceeds the 8 MB evidence-slice limit.")
    output.extend(header_bytes)
    stored_rows = 0
    for row in rows:
        encoded = _row_bytes(row, delimiter)
        if len(output) + len(encoded) > RESULT_EVIDENCE_SLICE_MAX_BYTES:
            break
        output.extend(encoded)
        stored_rows += 1
    return bytes(output), stored_rows


def _profile_small_result(path: Path) -> tuple[list[str], int | None, str]:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        header, _, row_count, delimiter = _csv_rows(path)
        return [str(value or "").strip() for value in header], row_count, (
            "tsv" if delimiter == "\t" else "csv"
        )
    if suffix == ".xlsx":
        header, _, row_count, sheet = _xlsx_rows(path)
        return [str(value or "").strip() for value in header], row_count, f"xlsx:{sheet}"
    return [], None, suffix.lstrip(".") or "binary"


def prepare_result_evidence(path: Path) -> RetainedResult:
    """Return the managed payload and retention metadata for one SQL result file."""

    path = path.resolve()
    if not path.is_file():
        raise ValueError(f"Result evidence file not found: {path}")
    source_size = path.stat().st_size
    source_sha = file_sha256(path)
    suffix = path.suffix.lower()
    sliced = source_size > RESULT_EVIDENCE_MAX_BYTES

    if sliced:
        if suffix not in SLICEABLE_RESULT_EXTENSIONS:
            raise ValueError(
                f"Result evidence over 10 MB must be CSV, TSV, or XLSX so it can be sliced; got {suffix or '(none)'}."
            )
        if suffix in {".csv", ".tsv"}:
            header, rows, row_count, delimiter = _csv_rows(path)
            source_format = "tsv" if delimiter == "\t" else "csv"
            stored_suffix = suffix
            stored_delimiter = delimiter
        else:
            header, rows, row_count, sheet = _xlsx_rows(path)
            source_format = f"xlsx:{sheet}"
            stored_suffix = ".csv"
            stored_delimiter = ","
        payload, stored_row_count = _render_slice(header, rows, stored_delimiter)
        sampling_method = (
            f"head_{RESULT_EVIDENCE_HEAD_ROWS}_tail_{RESULT_EVIDENCE_TAIL_ROWS}"
        )
        policy = "slice_over_10mb"
        full_source_stored = False
        columns = [str(value or "").strip() for value in header]
    else:
        payload = path.read_bytes()
        stored_suffix = suffix
        stored_row_count = None
        columns, row_count, source_format = _profile_small_result(path)
        sampling_method = "full_file"
        policy = "full_at_or_below_10mb"
        full_source_stored = True

    stored_sha = hashlib.sha256(payload).hexdigest()
    retention = {
        "contract_version": RESULT_EVIDENCE_RETENTION_VERSION,
        "payload_role": "sql_output_preview",
        "policy": policy,
        "threshold_bytes": RESULT_EVIDENCE_MAX_BYTES,
        "source_size_bytes": source_size,
        "source_sha256": source_sha,
        "source_file_name": path.name,
        "source_format": source_format,
        "source_row_count": row_count,
        "columns": columns,
        "is_sliced": sliced,
        "sampling_method": sampling_method,
        "stored_row_count": stored_row_count,
        "stored_size_bytes": len(payload),
        "stored_sha256": stored_sha,
        "stored_file_type": stored_suffix.lstrip("."),
        "full_source_stored_in_managed_assets": full_source_stored,
    }
    return RetainedResult(
        payload=payload,
        suffix=stored_suffix,
        media_type=mimetypes.guess_type(f"result{stored_suffix}")[0] or "application/octet-stream",
        stored_sha256=stored_sha,
        retention=retention,
    )


def copy_result_evidence(source: Path, destination: Path) -> tuple[Path, dict[str, Any]]:
    retained = prepare_result_evidence(source)
    actual_destination = destination.with_suffix(retained.suffix)
    write_retained_result(retained, actual_destination)
    return actual_destination, retained.retention


def write_retained_result(retained: RetainedResult, destination: Path) -> None:
    """Write and verify a payload already prepared by ``prepare_result_evidence``."""

    actual_destination = destination.with_suffix(retained.suffix)
    actual_destination.parent.mkdir(parents=True, exist_ok=True)
    actual_destination.write_bytes(retained.payload)
    if file_sha256(actual_destination) != retained.stored_sha256:
        actual_destination.unlink(missing_ok=True)
        raise ValueError("Stored result-evidence slice failed SHA-256 verification.")


def full_reusable_output_retention(path: Path, kind: str) -> dict[str, Any]:
    path = path.resolve()
    if kind not in REUSABLE_OUTPUT_KINDS:
        raise ValueError(f"Unsupported reusable output kind: {kind}")
    size = path.stat().st_size
    digest = file_sha256(path)
    return {
        "contract_version": RESULT_EVIDENCE_RETENTION_VERSION,
        "payload_role": "reusable_visual_or_analysis_asset",
        "policy": "full_reusable_output",
        "threshold_bytes": RESULT_EVIDENCE_MAX_BYTES,
        "source_size_bytes": size,
        "source_sha256": digest,
        "source_file_name": path.name,
        "source_format": path.suffix.lower().lstrip(".") or "binary",
        "source_row_count": None,
        "columns": [],
        "is_sliced": False,
        "sampling_method": "full_file",
        "stored_row_count": None,
        "stored_size_bytes": size,
        "stored_sha256": digest,
        "stored_file_type": path.suffix.lower().lstrip(".") or "binary",
        "full_source_stored_in_managed_assets": True,
    }


def copy_full_reusable_output(source: Path, destination: Path, kind: str) -> dict[str, Any]:
    retention = full_reusable_output_retention(source, kind)
    destination.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(source, destination)
    if file_sha256(destination) != retention["stored_sha256"]:
        destination.unlink(missing_ok=True)
        raise ValueError("Stored reusable output failed SHA-256 verification.")
    return retention
