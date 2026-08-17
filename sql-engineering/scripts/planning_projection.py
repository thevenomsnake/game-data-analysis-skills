#!/usr/bin/env python3
"""Export reviewed projections from exact planning-source releases.

The source is resolved through one project's active planning-source release.
Project-specific sheet and field choices live in repository projection specs.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
from datetime import datetime
from pathlib import Path
from pathlib import PurePosixPath
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string
import planning_source_provider as source_provider


SPEC_SCHEMA = "planning_projection_spec_v1"
RELEASE_SCHEMAS = {"planning_source_release_v1", "planning_source_release_v2"}
FILES_SCHEMA = "planning_source_files_v1"
BINDING_SCHEMAS = {"planning_source_binding_v1", "planning_source_binding_v2"}


class PlanningProjectionError(ValueError):
    pass


def assert_inside(path: Path, root: Path, label: str) -> Path:
    resolved = path.resolve()
    try:
        resolved.relative_to(root.resolve())
    except ValueError as error:
        raise PlanningProjectionError(f"{label} must stay inside {root}: {resolved}") from error
    return resolved


def safe_relative(value: str) -> str:
    relative = PurePosixPath(value.replace("\\", "/"))
    if relative.is_absolute() or not relative.parts or ".." in relative.parts:
        raise PlanningProjectionError(f"invalid planning source relative path: {value}")
    return relative.as_posix()


def read_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def write_json(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_header_text(value: Any) -> str:
    return clean_text(value).replace("\r\n", "\n").replace("\r", "\n")


def normalize_var(value: Any) -> str:
    return clean_text(value)


def base_var(value: Any) -> str:
    return normalize_var(value).split("#", 1)[0].strip()


def csv_value(value: Any) -> Any:
    if value is None:
        return ""
    return value


def sheet_values(ws) -> list[list[Any]]:
    return [list(row) for row in ws.iter_rows(values_only=True)]


def value_at(rows: list[list[Any]], row_idx_1based: int, col_idx_1based: int) -> Any:
    row_idx = row_idx_1based - 1
    col_idx = col_idx_1based - 1
    if row_idx < 0 or row_idx >= len(rows):
        return None
    row = rows[row_idx]
    if col_idx < 0 or col_idx >= len(row):
        return None
    return row[col_idx]


def make_headers(rows: list[list[Any]], header_rows: dict[str, int]) -> list[dict[str, Any]]:
    comment_row = int(header_rows["comment"])
    var_row = int(header_rows["var"])
    type_row = int(header_rows["type"])
    max_column = max((len(row) for row in rows), default=0)
    headers: list[dict[str, Any]] = []
    for col_idx in range(1, max_column + 1):
        headers.append(
            {
                "source_index": col_idx,
                "source_comment": normalize_header_text(value_at(rows, comment_row, col_idx)),
                "source_var": normalize_var(value_at(rows, var_row, col_idx)),
                "source_type": clean_text(value_at(rows, type_row, col_idx)),
            }
        )
    return headers


def find_column(headers: list[dict[str, Any]], keep: dict[str, Any]) -> dict[str, Any]:
    candidates = headers

    source_comment = keep.get("source_comment")
    if source_comment is not None:
        wanted_comment = normalize_header_text(source_comment)
        candidates = [h for h in candidates if h["source_comment"] == wanted_comment]

    source_var = keep.get("source_var")
    if source_var is not None:
        wanted_var = normalize_var(source_var)
        wanted_base = base_var(source_var)

        exact = [h for h in candidates if h["source_var"] == wanted_var]
        if exact:
            candidates = exact
        else:
            candidates = [h for h in candidates if base_var(h["source_var"]) == wanted_base]

    if not candidates:
        raise PlanningProjectionError(f"Column not found for keep spec: {keep}")

    if len(candidates) > 1:
        detail = [
            {
                "source_index": h["source_index"],
                "source_comment": h["source_comment"],
                "source_var": h["source_var"],
            }
            for h in candidates
        ]
        raise PlanningProjectionError(f"Ambiguous column for keep spec {keep}: {detail}")

    return candidates[0]


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def export_primary_sheet(
    wb,
    spec: dict[str, Any],
    sheet_spec: dict[str, Any],
    outputs_root: Path,
) -> dict[str, Any]:
    sheet_name = sheet_spec["sheet"]
    if sheet_name not in wb.sheetnames:
        raise PlanningProjectionError(f"Required sheet missing: {sheet_name}")

    ws = wb[sheet_name]
    rows_raw = sheet_values(ws)
    headers = make_headers(rows_raw, sheet_spec["header_rows"])

    selected: list[dict[str, Any]] = []
    used_indexes: set[int] = set()
    for keep in sheet_spec["keep_columns"]:
        header = find_column(headers, keep)
        output_name = keep["output"]
        selected_header = dict(header)
        selected_header["name"] = output_name
        if header["source_var"] != keep.get("source_var", header["source_var"]):
            selected_header["matched_from_spec_var"] = keep.get("source_var", "")
        if keep.get("source_comment") and header["source_var"].startswith("##"):
            selected_header["note"] = f"Matched by source_comment={keep['source_comment']}"
        else:
            selected_header["note"] = ""
        selected.append(selected_header)
        used_indexes.add(header["source_index"])

    fieldnames = [c["name"] for c in selected]
    primary_key = sheet_spec.get("primary_key", [])
    data_start = int(sheet_spec["header_rows"]["data_start"])
    rows: list[dict[str, Any]] = []
    empty_pk_rows = 0

    for row_idx in range(data_start, len(rows_raw) + 1):
        row: dict[str, Any] = {}
        for col in selected:
            row[col["name"]] = csv_value(value_at(rows_raw, row_idx, col["source_index"]))

        has_any_value = any(clean_text(v) for v in row.values())
        if not has_any_value:
            continue

        if primary_key:
            has_pk = all(clean_text(row.get(pk)) for pk in primary_key)
            if not has_pk:
                empty_pk_rows += 1
                continue

        rows.append(row)

    output_path = outputs_root / sheet_spec["output"]
    write_csv(output_path, fieldnames, rows)

    pk_values: list[str] = []
    duplicate_count = 0
    if primary_key:
        seen: set[str] = set()
        for row in rows:
            key = "\x1f".join(clean_text(row.get(pk)) for pk in primary_key)
            pk_values.append(key)
            if key in seen:
                duplicate_count += 1
            else:
                seen.add(key)

    excluded = [
        {
            "source_index": h["source_index"],
            "source_comment": h["source_comment"],
            "source_var": h["source_var"],
            "source_type": h["source_type"],
            "reason": sheet_spec.get("drop_columns_reason", ""),
        }
        for h in headers
        if h["source_index"] not in used_indexes
    ]

    schema = {
        "schema_version": spec.get("schema_version", 1),
        "table_id": spec["table_id"],
        "status": spec.get("status", ""),
        "source_reference": spec["source"],
        "primary_sheet": sheet_name,
        "header_rows": sheet_spec["header_rows"],
        "primary_key": primary_key,
        "columns": selected,
        "excluded_columns": excluded,
        "related_outputs": [],
    }

    return {
        "kind": "primary",
        "sheet": sheet_name,
        "output_path": output_path,
        "schema": schema,
        "row_count": len(rows),
        "column_count": len(fieldnames),
        "duplicate_pk_count": duplicate_count,
        "empty_pk_row_count": empty_pk_rows,
    }


def export_enum_sheet(wb, sheet_spec: dict[str, Any], outputs_root: Path) -> list[dict[str, Any]]:
    sheet_name = sheet_spec["sheet"]
    if sheet_name not in wb.sheetnames:
        raise PlanningProjectionError(f"Required sheet missing: {sheet_name}")

    ws = wb[sheet_name]
    rows_raw = sheet_values(ws)
    exported: list[dict[str, Any]] = []

    for enum_spec in sheet_spec["outputs"]:
        output_columns = enum_spec["columns"]
        source_indexes = [column_index_from_string(c) for c in enum_spec["source_columns"]]
        fill_down = set(enum_spec.get("fill_down", []))
        last_values = {name: "" for name in fill_down}
        rows: list[dict[str, Any]] = []

        for row_idx in range(int(enum_spec["start_row"]), len(rows_raw) + 1):
            row: dict[str, Any] = {}
            raw_by_output: dict[str, str] = {}
            for output_name, col_idx in zip(output_columns, source_indexes):
                raw = clean_text(value_at(rows_raw, row_idx, col_idx))
                raw_by_output[output_name] = raw
                if output_name in fill_down:
                    if raw:
                        last_values[output_name] = raw
                    row[output_name] = last_values[output_name]
                else:
                    row[output_name] = raw

            non_parent_values = [
                raw_by_output[name]
                for name in output_columns
                if name not in fill_down
            ]
            if not any(non_parent_values):
                continue
            if not any(clean_text(v) for v in row.values()):
                continue
            rows.append(row)

        output_path = outputs_root / enum_spec["name"]
        write_csv(output_path, output_columns, rows)
        exported.append(
            {
                "kind": "enum",
                "sheet": sheet_name,
                "output_path": output_path,
                "row_count": len(rows),
                "column_count": len(output_columns),
                "columns": output_columns,
            }
        )

    return exported


def resolve_source(repo_root: Path, spec: dict[str, Any]) -> tuple[Path, dict[str, Any]]:
    source = spec.get("source")
    if not isinstance(source, dict):
        raise PlanningProjectionError("projection spec requires a source object")
    project_id = str(source.get("project_id") or "").strip()
    relative_file = safe_relative(str(source.get("relative_file") or ""))
    project_root = assert_inside(repo_root / "sql-projects" / project_id, repo_root / "sql-projects", "project")
    binding_path = project_root / "planning" / "source_binding.json"
    binding = read_json(binding_path)
    if binding.get("contract_version") not in BINDING_SCHEMAS or binding.get("project_id") != project_id:
        raise PlanningProjectionError(f"invalid project planning-source binding: {binding_path}")
    release_manifest = assert_inside(
        project_root / str(binding.get("release_manifest") or ""),
        repo_root / "planning-sources",
        "release manifest",
    )
    release = read_json(release_manifest)
    if release.get("contract_version") not in RELEASE_SCHEMAS:
        raise PlanningProjectionError(f"invalid planning-source release: {release_manifest}")
    for field in ("product_id", "stage_id", "active_release_id", "tree_sha256"):
        release_field = "release_id" if field == "active_release_id" else field
        if binding.get(field) != release.get(release_field):
            raise PlanningProjectionError(f"planning-source binding mismatch for {field}")
    files_manifest = read_json(release_manifest.parent / str(release.get("files_manifest") or "files.json"))
    if files_manifest.get("contract_version") != FILES_SCHEMA:
        raise PlanningProjectionError("invalid planning-source files manifest")
    rows = files_manifest.get("files")
    if not isinstance(rows, list):
        raise PlanningProjectionError("planning-source files manifest requires files")
    matched = [row for row in rows if isinstance(row, dict) and row.get("relative_path") == relative_file]
    if len(matched) != 1:
        raise PlanningProjectionError(
            f"planning source file is not uniquely present in {release['release_id']}: {relative_file}"
        )
    source_kind = (
        "folder_snapshot"
        if release.get("contract_version") == "planning_source_release_v1"
        else str(release.get("source_kind") or "")
    )
    expected_hash = str(matched[0].get("sha256") or "")
    if source_kind == "svn_revision":
        control = release.get("source_control") if isinstance(release.get("source_control"), dict) else {}
        local_config_path = repo_root / ".local" / "planning-sources" / f"{project_id}.json"
        credential_ref = None
        if local_config_path.is_file():
            local_config = read_json(local_config_path)
            credential_ref = (
                local_config.get("credential_ref")
                if isinstance(local_config.get("credential_ref"), dict)
                else None
            )
        try:
            auth = source_provider.resolve_svn_auth(credential_ref)
            source_file = source_provider.materialize_svn_file(
                repo_root=repo_root,
                release_id=str(release["release_id"]),
                source_control=control,
                relative_file=relative_file,
                expected_sha256=expected_hash,
                expected_size=int(matched[0].get("size_bytes") or 0),
                auth=auth,
            )
        except source_provider.PlanningSourceProviderError as error:
            raise PlanningProjectionError(str(error)) from error
    elif source_kind == "folder_snapshot":
        files_root = release_manifest.parent / str(release.get("files_root") or "files")
        source_file = assert_inside(
            files_root / Path(*PurePosixPath(relative_file).parts),
            files_root,
            "planning source file",
        )
        if not source_file.is_file():
            raise PlanningProjectionError(f"planning source file does not exist: {relative_file}")
    else:
        raise PlanningProjectionError(f"unsupported planning-source release kind: {source_kind}")
    source_hash = sha256_file(source_file)
    if source_hash != expected_hash:
        raise PlanningProjectionError(f"planning source file hash mismatch: {relative_file}")
    reference = {
        "contract_version": "planning_source_reference_v1",
        "project_id": project_id,
        "product_id": release["product_id"],
        "stage_id": release["stage_id"],
        "release_id": release["release_id"],
        "release_tree_sha256": release["tree_sha256"],
        "release_manifest": release["release_manifest"],
        "source_kind": source_kind,
        "relative_file": relative_file,
        "file_sha256": source_hash,
    }
    control = release.get("source_control") if isinstance(release.get("source_control"), dict) else {}
    if source_kind == "svn_revision":
        reference["source_revision"] = int(control.get("revision") or 0)
        reference["source_repository_uuid"] = str(control.get("repository_uuid") or "")
    return source_file, reference


def export_from_spec(*, repo_root: Path, spec_path: Path, output_dir: Path) -> dict[str, Any]:
    repo_root = repo_root.resolve()
    spec_path = assert_inside(spec_path, repo_root, "projection spec")
    output_dir = assert_inside(output_dir, repo_root, "projection output")
    spec = read_json(spec_path)
    if spec.get("schema_version") != SPEC_SCHEMA or str(spec.get("status") or "").lower() != "active":
        raise PlanningProjectionError("projection spec must be active planning_projection_spec_v1")
    source_excel, source_reference = resolve_source(repo_root, spec)
    output_dir.mkdir(parents=True, exist_ok=True)
    resolved_spec = {**spec, "source": source_reference}

    workbook = load_workbook(source_excel, read_only=True, data_only=True)
    validation = spec.get("validation", {})
    for required_sheet in validation.get("require_sheets", []):
        if required_sheet not in workbook.sheetnames:
            raise PlanningProjectionError(f"Required sheet missing: {required_sheet}")

    source_hash = sha256_file(source_excel)
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    outputs: list[dict[str, Any]] = []
    primary_result: dict[str, Any] | None = None

    for sheet_spec in spec["sheets"]:
        role = sheet_spec.get("role")
        if role == "primary":
            primary_result = export_primary_sheet(workbook, resolved_spec, sheet_spec, output_dir)
            outputs.append(primary_result)
        elif role == "enum":
            outputs.extend(export_enum_sheet(workbook, sheet_spec, output_dir))
        else:
            raise PlanningProjectionError(f"Unsupported sheet role: {role}")
    workbook.close()

    if primary_result is None:
        raise PlanningProjectionError("Spec must contain one primary sheet.")

    related_outputs = [
        {
            "kind": item["kind"],
            "sheet": item["sheet"],
            "path": Path(item["output_path"]).name,
            "row_count": item["row_count"],
        }
        for item in outputs
        if item is not primary_result
    ]
    primary_result["schema"]["related_outputs"] = related_outputs

    table_id = spec["table_id"]
    schema_path = output_dir / f"{table_id}.schema.json"
    profile_path = output_dir / f"{table_id}.profile.json"

    write_json(schema_path, primary_result["schema"])

    enum_counts = {
        Path(item["output_path"]).stem + "_count": item["row_count"]
        for item in outputs
        if item["kind"] == "enum"
    }
    profile = {
        "generated_at": generated_at,
        "tool": "planning_projection.py",
        "schema_version": SPEC_SCHEMA,
        "table_id": table_id,
        "status": spec.get("status", ""),
        "source_reference": source_reference,
        "source_sha256": source_hash,
        "outputs_root": output_dir.relative_to(repo_root).as_posix(),
        "row_count": primary_result["row_count"],
        "column_count": primary_result["column_count"],
        "duplicate_id_count": primary_result["duplicate_pk_count"],
        "empty_id_count": primary_result["empty_pk_row_count"],
        **enum_counts,
    }
    write_json(profile_path, profile)

    outputs.append(
        {
            "kind": "schema",
            "sheet": primary_result["sheet"],
            "output_path": schema_path,
            "row_count": len(primary_result["schema"]["columns"]),
            "column_count": 0,
        }
    )
    outputs.append(
        {
            "kind": "profile",
            "sheet": primary_result["sheet"],
            "output_path": profile_path,
            "row_count": 1,
            "column_count": len(profile),
        }
    )

    expected_columns = validation.get("expected_primary_columns")
    if expected_columns is not None and primary_result["column_count"] != int(expected_columns):
        raise PlanningProjectionError(
            f"Primary column count mismatch: expected {expected_columns}, "
            f"got {primary_result['column_count']}"
        )
    if validation.get("primary_key_unique") and primary_result["duplicate_pk_count"]:
        raise PlanningProjectionError(f"Duplicate primary key count: {primary_result['duplicate_pk_count']}")
    if validation.get("primary_key_non_empty") and primary_result["empty_pk_row_count"]:
        raise PlanningProjectionError(f"Rows with content but empty primary key: {primary_result['empty_pk_row_count']}")

    return {
        "outputs": outputs,
        "source_file": source_excel,
        "source_reference": source_reference,
        "output_dir": output_dir,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--repo-root", type=Path, default=Path.cwd())
    parser.add_argument("--spec", required=True, type=Path)
    parser.add_argument("--output-dir", required=True, type=Path)
    args = parser.parse_args()
    repo_root = args.repo_root.resolve()
    spec_path = args.spec if args.spec.is_absolute() else repo_root / args.spec
    output_dir = args.output_dir if args.output_dir.is_absolute() else repo_root / args.output_dir
    try:
        result = export_from_spec(repo_root=repo_root, spec_path=spec_path, output_dir=output_dir)
        print(
            json.dumps(
                {
                    "status": "exported",
                    "source_reference": result["source_reference"],
                    "outputs": [
                        {
                            "kind": item["kind"],
                            "sheet": item["sheet"],
                            "path": Path(item["output_path"]).relative_to(repo_root).as_posix(),
                            "row_count": item["row_count"],
                        }
                        for item in result["outputs"]
                    ],
                },
                ensure_ascii=False,
                indent=2,
            )
        )
    except (PlanningProjectionError, OSError) as error:
        parser.exit(2, f"BLOCKED: {error}\n")


if __name__ == "__main__":
    main()
