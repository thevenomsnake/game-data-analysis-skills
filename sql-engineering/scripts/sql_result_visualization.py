#!/usr/bin/env python3
"""Bind a returned SQL result and a verified visual Excel to one exact SQL asset."""

from __future__ import annotations

import argparse
import copy
import csv
import hashlib
import html
import json
import math
import mimetypes
import re
import sys
import tempfile
import zipfile
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from asset_provenance import build_generation_provenance, now_iso  # noqa: E402
from capability_registry import command_function_ids  # noqa: E402
from function_gate import (  # noqa: E402
    FunctionGateError,
    add_function_gate_arguments,
    exit_with_gate_error,
    require_user_function_selection,
    require_user_request,
)
from formal_asset_repository import (  # noqa: E402
    apply_plan as apply_formal_asset_plan,
    list_packages as list_formal_asset_packages,
    load_package as load_formal_asset_package,
    plan_package as plan_formal_asset_package,
    validate_receipt as validate_formal_asset_receipt,
)
from result_evidence_retention import (  # noqa: E402
    file_sha256,
    full_reusable_output_retention,
    prepare_result_evidence,
)
from sql_facts import analyze_sql_file, execution_fingerprint, logic_fingerprint  # noqa: E402
from sql_project import manifest_path, read_json, rebuild_index  # noqa: E402
from sql_execution_adapter import effective_config_for_context  # noqa: E402
from sql_result_inspector import (  # noqa: E402
    inspect_result_file,
    time_coverage_problem_messages,
)
from sql_summary_planner import load_analysis_bundle  # noqa: E402
from sql_query_workspace import (  # noqa: E402
    INDEX_REL,
    _index_files,
    _write_transaction,
    attach_derived_output,
    find_query_reference,
    json_text,
    load_index,
    transition_query,
)
from workbook_manifest import build_workbook_manifest, xlsx_chart_parts  # noqa: E402


RECEIPT_VERSION = "sql_result_visualization_receipt_v1"
BUNDLE_RECEIPT_VERSION = "sql_result_visualization_bundle_receipt_v1"
BUNDLE_RESULT_RECEIPT_VERSION = "query_analysis_bundle_result_receipt_v1"
SOURCE_VERSION = "sql_result_source_v1"
BINDING_VERSION = "sql_result_binding_v1"
BUNDLE_OUTPUT_REF_VERSION = "query_analysis_bundle_output_ref_v1"
VALUE_REFRESH_VERSION = "visualization_value_refresh_v1"
RESULT_EXTENSIONS = {".csv", ".xlsx"}
FORMAL_QUERY_ROLES = {
    "formal_query",
    "formal_query_unverified",
    "formal_query_sql",
    "query_sql",
    "historical_query_sql",
}
FORMAL_DASHBOARD_ROLES = {"dashboard_delivery", "dashboard_delivery_sql", "dashboard_sql"}
MAX_VALUE_REFRESH_CELLS = 250_000
TOKENS_PATH = SCRIPT_DIR.parent / "assets" / "viz_tokens.json"
CHART_NS = {
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "c": "http://schemas.openxmlformats.org/drawingml/2006/chart",
}
DRAWING_NS = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
MAX_CONDITIONAL_FORMAT_AUDITS = 200
MAX_CONTRAST_VIOLATIONS = 20


def _hex_rgb(value: Any) -> str | None:
    text = re.sub(r"[^0-9A-Fa-f]", "", str(value or ""))
    if len(text) == 8:
        text = text[-6:]
    return text.upper() if len(text) == 6 else None


def _relative_luminance(rgb: str) -> float:
    channels: list[float] = []
    for offset in (0, 2, 4):
        channel = int(rgb[offset : offset + 2], 16) / 255.0
        channels.append(
            channel / 12.92
            if channel <= 0.04045
            else ((channel + 0.055) / 1.055) ** 2.4
        )
    return 0.2126 * channels[0] + 0.7152 * channels[1] + 0.0722 * channels[2]


def _contrast_ratio(foreground: str, background: str) -> float:
    foreground_luminance = _relative_luminance(foreground)
    background_luminance = _relative_luminance(background)
    lighter = max(foreground_luminance, background_luminance)
    darker = min(foreground_luminance, background_luminance)
    return (lighter + 0.05) / (darker + 0.05)


def _validate_color_scale_token(name: str, token: dict[str, Any]) -> None:
    foreground = _hex_rgb(token.get("foreground"))
    minimum = float(token.get("minimum_contrast_ratio") or 0)
    if not foreground or minimum < 1:
        raise RuntimeError(
            f"Visualization token `{name}` must define a foreground and minimum_contrast_ratio."
        )
    for endpoint_name in ("low", "mid", "high"):
        endpoint = _hex_rgb(token.get(endpoint_name))
        if not endpoint:
            raise RuntimeError(f"Visualization token `{name}.{endpoint_name}` must be an RGB color.")
        ratio = _contrast_ratio(foreground, endpoint)
        if ratio + 1e-9 < minimum:
            raise RuntimeError(
                f"Visualization token `{name}.{endpoint_name}` has {ratio:.2f}:1 contrast; "
                f"minimum is {minimum:.2f}:1."
            )


def _load_viz_tokens() -> dict[str, Any]:
    try:
        tokens = json.loads(TOKENS_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise RuntimeError(f"Visualization tokens are unreadable: {TOKENS_PATH}: {exc}") from exc
    if tokens.get("schema_version") != "viz_tokens_v3":
        raise RuntimeError("Visualization tokens must use schema_version `viz_tokens_v3`.")
    evidence = tokens.get("evidence") or {}
    if not evidence.get("base_field_terms") or not evidence.get("base_display_terms"):
        raise RuntimeError("Visualization tokens must define Base field and display terms.")
    chart_series = tokens.get("chart_series") or {}
    peer_sequence = list(chart_series.get("peer_sequence") or [])
    if not 2 <= len(peer_sequence) <= int(chart_series.get("max_equal_weight_colors") or 0):
        raise RuntimeError("Visualization tokens must define two to four governed peer-series colors.")
    if len({_hex_rgb(color) for color in peer_sequence}) != len(peer_sequence) or any(
        _hex_rgb(color) is None for color in peer_sequence
    ):
        raise RuntimeError("Visualization peer-series colors must be unique RGB values.")
    conditional = tokens.get("conditional_formatting") or {}
    if (conditional.get("percentage_color_scale") or {}).get("kind") != "color_scale":
        raise RuntimeError("Visualization tokens must define a percentage color scale.")
    for token_name in ("percentage_color_scale", "signed_percentage_color_scale"):
        token = conditional.get(token_name) or {}
        if token.get("kind") != "color_scale":
            raise RuntimeError(f"Visualization tokens must define `{token_name}` as a color scale.")
        _validate_color_scale_token(token_name, token)
    if (conditional.get("numeric_data_bar") or {}).get("kind") != "data_bar":
        raise RuntimeError("Visualization tokens must define a numeric data bar.")
    return tokens


VIZ_TOKENS = _load_viz_tokens()


def _format_percentile_point(value: float) -> str:
    rounded = round(value, 6)
    text = str(int(rounded)) if float(rounded).is_integer() else f"{rounded:g}"
    return f"P{text}"


def _percentile_points(values: list[str]) -> list[str]:
    points: set[str] = set()
    for value in values:
        text = str(value or "").strip().casefold()
        if not text:
            continue
        if re.search(r"(?:^|[^a-z])(median)(?:$|[^a-z])|中位", text):
            points.add("P50")
        for match in re.finditer(r"(?<![a-z])p\s*([0-9]{1,3}(?:\.[0-9]+)?)(?![a-z0-9])", text):
            number = float(match.group(1))
            if 0 <= number <= 100:
                points.add(_format_percentile_point(number))
        for match in re.finditer(
            r"(?:percentile|quantile)[_\s-]*(0?\.\d+|[0-9]{1,3}(?:\.[0-9]+)?)",
            text,
        ):
            number = float(match.group(1))
            if 0 < number <= 1:
                number *= 100
            if 0 <= number <= 100:
                points.add(_format_percentile_point(number))
        for match in re.finditer(r"([0-9]{1,3}(?:\.[0-9]+)?)\s*%?\s*(?:百分位|分位)", text):
            number = float(match.group(1))
            if 0 <= number <= 100:
                points.add(_format_percentile_point(number))
        for match in re.finditer(r"(?<![a-z0-9])q([123])(?![a-z0-9])", text):
            points.add({"1": "P25", "2": "P50", "3": "P75"}[match.group(1)])
    return sorted(points, key=lambda item: float(item[1:]))


def _label_key(value: Any) -> str:
    return re.sub(r"[\s`\"'\[\]_:：/\\()（）-]+", "", str(value or "")).casefold()


def _metric_stem(value: Any) -> str:
    stem = _label_key(value)
    suffixes = (
        "percentage",
        "conversionrate",
        "retentionrate",
        "ratio",
        "share",
        "占比",
        "百分比",
        "比例",
        "比率",
        "转化率",
        "留存率",
        "玩家数",
        "用户数",
        "角色数",
        "设备数",
        "人数",
        "人次",
        "数量",
        "count",
        "率",
        "数",
    )
    changed = True
    while stem and changed:
        changed = False
        for suffix in suffixes:
            if stem.endswith(suffix) and len(stem) > len(suffix):
                stem = stem[: -len(suffix)]
                changed = True
                break
    for entity in ("玩家", "用户", "角色", "设备", "事件"):
        if stem.endswith(entity) and len(stem) > len(entity):
            stem = stem[: -len(entity)]
            break
    return stem


def _base_field_candidates(result_inspections: list[dict[str, Any]]) -> list[str]:
    terms = [_label_key(item) for item in VIZ_TOKENS["evidence"]["base_field_terms"]]
    ratio_field_keys = {
        _label_key(rule.get("output_field"))
        for inspection in result_inspections
        for rule in inspection.get("ratio_field_rules", [])
        if rule.get("output_field")
    }
    candidates: list[str] = []
    for inspection in result_inspections:
        for column in inspection.get("columns", []):
            label = str(column or "").strip()
            key = _label_key(label)
            if (
                label
                and key not in ratio_field_keys
                and any(term and term in key for term in terms)
                and label not in candidates
            ):
                candidates.append(label)
    explicit = [
        label
        for label in candidates
        if re.search(
            r"(?:base|denominator|sample|分母|样本|^(?:总|总体|整体)(?:玩家|用户|角色|设备)?(?:数|人数)$)",
            _label_key(label),
            flags=re.IGNORECASE,
        )
    ]
    return explicit or candidates


def _base_field_values(
    result_inspections: list[dict[str, Any]],
    fields: list[str],
) -> dict[str, list[str]]:
    values: dict[str, list[str]] = {field: [] for field in fields}
    for inspection in result_inspections:
        for row in inspection.get("sample_rows", []):
            if not isinstance(row, dict):
                continue
            for field in fields:
                value = str(row.get(field) or "").strip()
                if value and value not in values[field]:
                    values[field].append(value)
    return values


def _base_requirement_reasons(
    result_inspections: list[dict[str, Any]],
    chart_audits: list[dict[str, Any]],
) -> tuple[list[str], list[str]]:
    ratio_fields = list(
        dict.fromkeys(
            str(rule.get("output_field") or "").strip()
            for inspection in result_inspections
            for rule in inspection.get("ratio_field_rules", [])
            if str(rule.get("output_field") or "").strip()
        )
    )
    reasons: list[str] = []
    if ratio_fields:
        reasons.append("ratio_fields")
    if any(audit.get("normalized_compositions") for audit in chart_audits):
        reasons.append("normalized_composition")
    return reasons, ratio_fields


def _redundant_absolute_fields(
    result_inspections: list[dict[str, Any]],
    ratio_fields: list[str],
) -> list[str]:
    ratio_keys = {_label_key(field) for field in ratio_fields}
    ratio_stems = {_metric_stem(field) for field in ratio_fields if _metric_stem(field)}
    count_suffix = re.compile(r"(?:玩家数|用户数|角色数|设备数|人数|人次|数量|count|数)$", re.IGNORECASE)
    fields: list[str] = []
    for inspection in result_inspections:
        for column in inspection.get("columns", []):
            label = str(column or "").strip()
            if (
                label
                and _label_key(label) not in ratio_keys
                and count_suffix.search(_label_key(label))
                and _metric_stem(label) in ratio_stems
                and label not in fields
            ):
                fields.append(label)
    return fields


def _has_values_below(sheet: Any, cell: Any) -> bool:
    return any(
        sheet.cell(row=row_number, column=cell.column).value is not None
        for row_number in range(cell.row + 1, sheet.max_row + 1)
    )


def _base_display_layout(
    presentation_sheets: list[Any],
    *,
    fields: list[str],
    ratio_fields: list[str],
    redundant_absolute_fields: list[str],
) -> dict[str, Any]:
    field_keys = {_label_key(field): field for field in fields}
    ratio_keys = {_label_key(field): field for field in ratio_fields}
    absolute_keys = {_label_key(field): field for field in redundant_absolute_fields}
    header_blocks: list[dict[str, Any]] = []
    observations: list[dict[str, Any]] = []

    for sheet in presentation_sheets:
        for row_number in range(1, sheet.max_row + 1):
            row_cells = [sheet.cell(row=row_number, column=column) for column in range(1, sheet.max_column + 1)]
            runs: list[list[Any]] = []
            current: list[Any] = []
            for cell in row_cells:
                if cell.value is not None and str(cell.value).strip():
                    current.append(cell)
                elif current:
                    runs.append(current)
                    current = []
            if current:
                runs.append(current)

            for run in runs:
                bases = [
                    {"field": field_keys[_label_key(cell.value)], "cell": cell.coordinate, "column": cell.column}
                    for cell in run
                    if _label_key(cell.value) in field_keys and _has_values_below(sheet, cell)
                ]
                ratios = [
                    {"field": ratio_keys[_label_key(cell.value)], "cell": cell.coordinate, "column": cell.column}
                    for cell in run
                    if _label_key(cell.value) in ratio_keys and _has_values_below(sheet, cell)
                ]
                absolutes = [
                    {"field": absolute_keys[_label_key(cell.value)], "cell": cell.coordinate, "column": cell.column}
                    for cell in run
                    if _label_key(cell.value) in absolute_keys and _has_values_below(sheet, cell)
                ]
                if not bases or not ratios:
                    continue
                block = {
                    "sheet": sheet.title,
                    "header_row": row_number,
                    "range": f"{run[0].coordinate}:{run[-1].coordinate}",
                    "base_fields": [item["field"] for item in bases],
                    "ratio_fields": [item["field"] for item in ratios],
                    "redundant_absolute_fields": [item["field"] for item in absolutes],
                }
                header_blocks.append(block)

                if len(bases) > 1:
                    observations.append(
                        {
                            **block,
                            "reason": "multiple_base_columns",
                            "fields": [item["field"] for item in bases],
                        }
                    )
                if max(item["column"] for item in bases) > min(item["column"] for item in ratios):
                    observations.append(
                        {
                            **block,
                            "reason": "base_after_ratio",
                            "fields": [item["field"] for item in bases],
                        }
                    )
                total_penetration = [
                    item
                    for item in ratios
                    if re.search(r"(?:总|整体|总体)(?:渗透|覆盖)", item["field"])
                ]
                if total_penetration and min(item["column"] for item in ratios) != total_penetration[0]["column"]:
                    observations.append(
                        {
                            **block,
                            "reason": "total_penetration_not_first",
                            "fields": [item["field"] for item in total_penetration],
                        }
                    )
                if absolutes:
                    observations.append(
                        {
                            **block,
                            "reason": "redundant_absolute_with_ratio",
                            "fields": [item["field"] for item in absolutes],
                        }
                    )

    return {
        "policy_version": "base_display_layout_v1",
        "header_blocks": header_blocks[:100],
        "header_blocks_truncated": len(header_blocks) > 100,
        "observations": observations[:20],
        "observation_count": len(observations),
        "review_recommended": bool(observations),
        "complete": True,
    }


def _base_coverage(
    path: Path,
    *,
    result_inspections: list[dict[str, Any]],
    chart_audits: list[dict[str, Any]],
) -> dict[str, Any]:
    reasons, ratio_fields = _base_requirement_reasons(result_inspections, chart_audits)
    required = bool(result_inspections and reasons)
    fields = _base_field_candidates(result_inspections) if required else []
    redundant_absolute_fields = _redundant_absolute_fields(result_inspections, ratio_fields)
    field_values = _base_field_values(result_inspections, fields)
    coverage = {
        "required": required,
        "reasons": reasons,
        "ratio_fields": ratio_fields,
        "available_fields": fields,
        "presentation_sheets": [],
        "visible_fields": [],
        "visible_base_texts": [],
        "display_mode": "not_required" if not required else "missing",
        "complete": not required,
        "display_layout": {
            "policy_version": "base_display_layout_v1",
            "header_blocks": [],
            "header_blocks_truncated": False,
            "observations": [],
            "observation_count": 0,
            "review_recommended": False,
            "complete": True,
        },
        "manual_checks": list(VIZ_TOKENS["evidence"].get("manual_checks") or []),
    }
    if not required:
        return coverage
    if not fields:
        raise ValueError(
            "[VIS-BASE-001] Ratio/distribution result lacks an explicit Base/denominator field. "
            "Return the same-grain sample/player/entity count from QUERY before visual finalization."
        )

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - runtime dependency guard
        raise ValueError("[VIS-BASE-003] Base presentation QA requires openpyxl.") from exc

    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        presentation_sheets = [sheet for sheet in workbook.worksheets if sheet._charts]
        coverage["presentation_sheets"] = [sheet.title for sheet in presentation_sheets]
        visible_fields: list[str] = []
        field_keys = {_label_key(field): field for field in fields}
        for sheet in presentation_sheets:
            for row in sheet.iter_rows():
                for cell in row:
                    field = field_keys.get(_label_key(cell.value))
                    if not field or field in visible_fields:
                        continue
                    if _has_values_below(sheet, cell):
                        visible_fields.append(field)
        coverage["display_layout"] = _base_display_layout(
            presentation_sheets,
            fields=fields,
            ratio_fields=ratio_fields,
            redundant_absolute_fields=redundant_absolute_fields,
        )
        cell_texts = [
            str(cell.value).strip()
            for sheet in presentation_sheets
            for row in sheet.iter_rows()
            for cell in row
            if cell.value is not None and str(cell.value).strip()
        ]
    finally:
        workbook.close()

    display_terms = [_label_key(item) for item in VIZ_TOKENS["evidence"]["base_display_terms"]]
    visible_base_texts = [
        value
        for value in cell_texts
        if any(term and term in _label_key(value) for term in display_terms)
        and bool(re.search(r"\d", value))
    ]
    common_value_texts = {
        re.sub(r"[,\s]", "", values[0]).casefold()
        for values in field_values.values()
        if len(values) == 1
    }
    common_subtitle = any(
        any(token and token in re.sub(r"[,\s]", "", text).casefold() for token in common_value_texts)
        for text in visible_base_texts
    )
    coverage["visible_fields"] = visible_fields
    coverage["visible_base_texts"] = visible_base_texts[:8]
    if visible_fields:
        coverage["display_mode"] = "adjacent_table_or_label"
        coverage["complete"] = True
    elif common_subtitle:
        coverage["display_mode"] = "common_base_subtitle"
        coverage["complete"] = True
    if not coverage["complete"]:
        raise ValueError(
            "[VIS-BASE-002] Base exists in the result but is not visible on a chart-containing presentation sheet. "
            "Show same-grain Base as an adjacent column/label, or show one proven common Base in the chart subtitle. "
            "A source-data sheet or distant note does not count."
        )
    return coverage


def _is_summary_label(value: Any) -> bool:
    text = re.sub(r"\s+", "", str(value or ""))
    if not text:
        return False
    return bool(
        re.fullmatch(
            r"(?:合计|总计|汇总|整体(?:平均|均值|转化率|留存率|占比|比率|P\d+(?:\.\d+)?)|加权平均)",
            text,
            flags=re.IGNORECASE,
        )
    )


def _merge_audit(
    path: Path,
    *,
    result_inspections: list[dict[str, Any]],
) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - runtime dependency guard
        raise ValueError("[VIS-LAYOUT-001] Merge QA requires openpyxl.") from exc

    result_field_keys = {
        _label_key(column)
        for inspection in result_inspections
        for column in inspection.get("columns", [])
        if str(column or "").strip()
    }
    allowed: list[dict[str, str]] = []
    blocked: list[dict[str, str]] = []
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        for sheet in workbook.worksheets:
            is_presentation = bool(sheet._charts)
            for merged_range in sheet.merged_cells.ranges:
                cell_range = str(merged_range)
                reasons: list[str] = []
                role = ""
                top_left = sheet.cell(merged_range.min_row, merged_range.min_col).value
                if not is_presentation:
                    reasons.append("source or non-presentation sheets must remain unmerged")
                if top_left is None or not str(top_left).strip():
                    reasons.append("merged group label is empty")
                if _label_key(top_left) in result_field_keys:
                    reasons.append("result/header fields must remain independently selectable")
                if _is_summary_label(top_left):
                    reasons.append("summary labels must remain on an independent row")

                is_vertical_group = (
                    merged_range.min_col == merged_range.max_col
                    and merged_range.min_row < merged_range.max_row
                )
                is_horizontal_header = (
                    merged_range.min_row == merged_range.max_row
                    and merged_range.min_col < merged_range.max_col
                )
                if is_vertical_group:
                    role = "vertical_table_parent_group"
                    for row_number in range(merged_range.min_row, merged_range.max_row + 1):
                        detail_values = [
                            sheet.cell(row_number, column_number).value
                            for column_number in range(merged_range.max_col + 1, sheet.max_column + 1)
                            if sheet.cell(row_number, column_number).value is not None
                            and str(sheet.cell(row_number, column_number).value).strip()
                        ]
                        if not detail_values:
                            reasons.append(f"row {row_number} has no populated child/detail cells to the right")
                            continue
                        if any(_is_summary_label(value) for value in detail_values):
                            reasons.append(f"row {row_number} is a summary row and cannot be merged into a parent group")
                elif is_horizontal_header:
                    role = "horizontal_table_header_group"
                    same_row_outside = [
                        sheet.cell(merged_range.min_row, column_number).value
                        for column_number in range(1, sheet.max_column + 1)
                        if not (merged_range.min_col <= column_number <= merged_range.max_col)
                        and sheet.cell(merged_range.min_row, column_number).value is not None
                        and str(sheet.cell(merged_range.min_row, column_number).value).strip()
                    ]
                    if not same_row_outside:
                        reasons.append("merged header occupies the whole populated row and behaves like a wide title")
                    child_row = merged_range.max_row + 1
                    child_headers = [
                        sheet.cell(child_row, column_number).value
                        for column_number in range(merged_range.min_col, merged_range.max_col + 1)
                    ] if child_row <= sheet.max_row else []
                    if not child_headers or any(
                        value is None or not str(value).strip()
                        for value in child_headers
                    ):
                        reasons.append("merged header is not immediately above one populated child header per column")
                    if any(_is_summary_label(value) for value in same_row_outside):
                        reasons.append("summary labels cannot participate in a merged table header")
                else:
                    reasons.append("multi-row multi-column merges are not a supported table grouping")

                if reasons:
                    blocked.append(
                        {
                            "sheet": sheet.title,
                            "range": cell_range,
                            "reason": "; ".join(dict.fromkeys(reasons)),
                        }
                    )
                else:
                    allowed.append(
                        {
                            "sheet": sheet.title,
                            "range": cell_range,
                            "role": role,
                        }
                    )
    finally:
        workbook.close()

    return {
        "policy_version": "table_semantic_merge_v1",
        "allowed_merges": allowed,
        "blocked_merges": blocked,
        "manual_checks": [
            "Each allowed vertical merge must represent one exact repeated parent-dimension value across contiguous child rows.",
            "Each allowed horizontal merge must be a nested table header immediately above its populated child headers.",
            "The merge must stay inside the intended copied table footprint and must not cross an unrelated group.",
        ],
    }


def _single_line_context_audit(path: Path) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - runtime dependency guard
        raise ValueError("[VIS-LAYOUT-002] Presentation text-layout QA requires openpyxl.") from exc

    metadata_label_keys = {
        _label_key(value)
        for value in (
            "统计窗口",
            "观察窗口",
            "新增窗口",
            "分析人群",
            "比较轴",
            "合计规则",
            "SQL版本",
            "来源",
            "数据截至",
            "统计周期",
        )
    }
    evidence_label_keys = {
        _label_key(value)
        for value in ("读图", "质量说明", "排除说明", "结果Base", "口径提示")
    }
    label_keys = metadata_label_keys | evidence_label_keys | {_label_key("单位"), _label_key("粒度")}
    context_rows: list[dict[str, Any]] = []
    violations: list[dict[str, Any]] = []
    observations: list[dict[str, Any]] = []
    context_row_count = 0
    violation_count = 0
    observation_count = 0
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        for sheet in (item for item in workbook.worksheets if item._charts):
            for row_number in range(1, sheet.max_row + 1):
                cells = [
                    sheet.cell(row=row_number, column=column)
                    for column in range(1, sheet.max_column + 1)
                    if sheet.cell(row=row_number, column=column).value is not None
                    and str(sheet.cell(row=row_number, column=column).value).strip()
                ]
                metadata_labels = [
                    str(cell.value).strip()
                    for cell in cells
                    if _label_key(cell.value) in metadata_label_keys
                ]
                evidence_labels = [
                    str(cell.value).strip()
                    for cell in cells
                    if _label_key(cell.value) in evidence_label_keys
                ]
                if not metadata_labels and not evidence_labels:
                    continue
                labels = [
                    str(cell.value).strip()
                    for cell in cells
                    if _label_key(cell.value) in label_keys
                ]
                context_row_count += 1
                wrapped_cells = [
                    cell.coordinate
                    for cell in cells
                    if cell.alignment.wrap_text is True or "\n" in str(cell.value)
                ]
                record = {
                    "sheet": sheet.title,
                    "row": row_number,
                    "labels": labels,
                    "context_kind": "metadata" if metadata_labels else "evidence",
                    "wrapped_cells": wrapped_cells,
                }
                if len(context_rows) < 100:
                    context_rows.append(record)
                if wrapped_cells and metadata_labels:
                    violation_count += 1
                    if len(violations) < 20:
                        violations.append(record)
                elif wrapped_cells:
                    observation_count += 1
                    if len(observations) < 20:
                        observations.append(record)
    finally:
        workbook.close()

    return {
        "policy_version": "single_line_presentation_context_v1",
        "context_rows": context_rows,
        "context_rows_truncated": context_row_count > len(context_rows),
        "violations": violations,
        "violation_count": violation_count,
        "observations": observations,
        "observation_count": observation_count,
        "review_recommended": bool(observations),
        "complete": violation_count == 0,
    }


def _apply_excel_tint(rgb: str, tint: float) -> str:
    if abs(tint) <= 1e-9:
        return rgb
    channels: list[int] = []
    for offset in (0, 2, 4):
        value = int(rgb[offset : offset + 2], 16)
        adjusted = value * (1.0 + tint) if tint < 0 else value + (255 - value) * tint
        channels.append(max(0, min(255, round(adjusted))))
    return "".join(f"{channel:02X}" for channel in channels)


def _theme_color_map(workbook: Any) -> dict[int, str]:
    raw_theme = workbook.loaded_theme
    if not raw_theme:
        return {}
    try:
        root = ET.fromstring(raw_theme)
    except ET.ParseError:
        return {}
    scheme = root.find(".//a:clrScheme", DRAWING_NS)
    if scheme is None:
        return {}
    colors: dict[int, str] = {}
    for index, slot in enumerate(list(scheme)):
        color_node = next(iter(slot), None)
        if color_node is None:
            continue
        rgb = _hex_rgb(color_node.attrib.get("val") or color_node.attrib.get("lastClr"))
        if rgb:
            colors[index] = rgb
    return colors


def _resolve_workbook_color(color: Any, theme_colors: dict[int, str]) -> str | None:
    if color is None:
        return None
    color_type = str(getattr(color, "type", "") or "")
    rgb: str | None = None
    if color_type == "rgb":
        rgb = _hex_rgb(getattr(color, "rgb", None))
    elif color_type == "theme":
        theme_index = getattr(color, "theme", None)
        if isinstance(theme_index, int):
            rgb = theme_colors.get(theme_index)
    elif color_type == "indexed":
        indexed = getattr(color, "indexed", None)
        if isinstance(indexed, int):
            try:
                from openpyxl.styles.colors import COLOR_INDEX

                if 0 <= indexed < len(COLOR_INDEX):
                    rgb = _hex_rgb(COLOR_INDEX[indexed])
            except ImportError:  # pragma: no cover - dependency is already required by caller
                return None
    if not rgb:
        return None
    tint = getattr(color, "tint", 0.0)
    try:
        return _apply_excel_tint(rgb, float(tint or 0.0))
    except (TypeError, ValueError):
        return rgb


def _cells_inside_range(sheet: Any, cell_range: Any) -> list[Any]:
    return [
        cell
        for (row, column), cell in sheet._cells.items()
        if cell_range.min_row <= row <= cell_range.max_row
        and cell_range.min_col <= column <= cell_range.max_col
        and cell.value is not None
    ]


def _conditional_format_contrast_audit(path: Path) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - runtime dependency guard
        raise ValueError("[VIS-CONTRAST-001] Conditional-format contrast QA requires openpyxl.") from exc

    token = VIZ_TOKENS["conditional_formatting"]["percentage_color_scale"]
    default_foreground = _hex_rgb(token.get("foreground")) or "17212B"
    minimum_ratio = float(token.get("minimum_contrast_ratio") or 4.5)
    workbook = load_workbook(path, read_only=False, data_only=False)
    audits: list[dict[str, Any]] = []
    violations: list[dict[str, Any]] = []
    violation_count = 0
    color_scale_rule_count = 0
    range_count = 0
    try:
        theme_colors = _theme_color_map(workbook)
        for sheet in workbook.worksheets:
            for conditional in sheet.conditional_formatting:
                for rule in sheet.conditional_formatting[conditional]:
                    if str(getattr(rule, "type", "") or "") != "colorScale":
                        continue
                    color_scale_rule_count += 1
                    color_scale = getattr(rule, "colorScale", None)
                    raw_endpoints = list(getattr(color_scale, "color", []) or [])
                    endpoint_colors = [
                        _resolve_workbook_color(color, theme_colors)
                        for color in raw_endpoints
                    ]
                    for cell_range in conditional.sqref.ranges:
                        range_count += 1
                        cells = _cells_inside_range(sheet, cell_range)
                        if not cells:
                            audit = {
                                "sheet": sheet.title,
                                "range": str(cell_range),
                                "semantic_kind": "color_scale",
                                "foreground_colors": [],
                                "endpoint_colors": [
                                    f"#{color}" if color else "unresolved"
                                    for color in endpoint_colors
                                ],
                                "minimum_observed_contrast": None,
                                "status": "skipped_empty",
                            }
                            if len(audits) < MAX_CONDITIONAL_FORMAT_AUDITS:
                                audits.append(audit)
                            continue

                        foregrounds: dict[str, str] = {}
                        unresolved_foreground = False
                        percentage_cells = 0
                        for cell in cells:
                            font_color = cell.font.color
                            foreground = _resolve_workbook_color(font_color, theme_colors)
                            if not foreground and font_color is not None:
                                unresolved_foreground = True
                                continue
                            foreground = foreground or default_foreground
                            foregrounds.setdefault(foreground, cell.coordinate)
                            if "%" in str(cell.number_format or ""):
                                percentage_cells += 1

                        ratios: list[tuple[float, str, str]] = []
                        unresolved_endpoint = any(color is None for color in endpoint_colors)
                        for foreground in foregrounds:
                            for endpoint in endpoint_colors:
                                if endpoint:
                                    ratios.append(
                                        (_contrast_ratio(foreground, endpoint), foreground, endpoint)
                                    )
                        worst = min(ratios, default=None, key=lambda item: item[0])
                        status = "pass"
                        if unresolved_endpoint or unresolved_foreground or not endpoint_colors:
                            status = "unresolved"
                        elif worst and worst[0] + 1e-9 < minimum_ratio:
                            status = "blocked"
                        audit = {
                            "sheet": sheet.title,
                            "range": str(cell_range),
                            "semantic_kind": "percentage_or_rate" if percentage_cells else "color_scale",
                            "foreground_colors": [f"#{color}" for color in sorted(foregrounds)]
                            + (["unresolved"] if unresolved_foreground else []),
                            "endpoint_colors": [
                                f"#{color}" if color else "unresolved"
                                for color in endpoint_colors
                            ],
                            "minimum_observed_contrast": round(worst[0], 3) if worst else None,
                            "status": status,
                        }
                        if len(audits) < MAX_CONDITIONAL_FORMAT_AUDITS:
                            audits.append(audit)
                        if status != "pass":
                            violation_count += 1
                            violation = {
                                **audit,
                                "sample_cell": (
                                    foregrounds.get(worst[1], cells[0].coordinate)
                                    if worst
                                    else cells[0].coordinate
                                ),
                                "minimum_required_contrast": minimum_ratio,
                            }
                            if worst:
                                violation["worst_foreground"] = f"#{worst[1]}"
                                violation["worst_background"] = f"#{worst[2]}"
                            if len(violations) < MAX_CONTRAST_VIOLATIONS:
                                violations.append(violation)
    finally:
        workbook.close()

    return {
        "policy_version": "conditional_format_text_contrast_v1",
        "minimum_contrast_ratio": minimum_ratio,
        "color_scale_rule_count": color_scale_rule_count,
        "audited_range_count": range_count,
        "audits": audits,
        "audits_truncated": range_count > len(audits),
        "violation_count": violation_count,
        "violations": violations,
        "violations_truncated": violation_count > len(violations),
        "complete": violation_count == 0,
    }


def _chart_label_values(root: ET.Element) -> tuple[list[str], list[str]]:
    values: list[str] = []
    formulas: list[str] = []
    for series in root.findall(".//c:ser", CHART_NS):
        for container_name in ("c:tx", "c:cat"):
            container = series.find(container_name, CHART_NS)
            if container is None:
                continue
            values.extend(
                str(node.text or "").strip()
                for node in container.findall(".//a:t", CHART_NS) + container.findall(".//c:v", CHART_NS)
                if str(node.text or "").strip()
            )
            formulas.extend(
                str(node.text or "").strip()
                for node in container.findall(".//c:f", CHART_NS)
                if str(node.text or "").strip()
            )
    return values, formulas


def _resolve_chart_label_formulas(path: Path, formulas: list[str]) -> list[str]:
    if not formulas:
        return []
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - runtime dependency guard
        raise ValueError("Percentile chart QA requires openpyxl to resolve chart source labels.") from exc

    workbook = load_workbook(path, read_only=True, data_only=False)
    values: list[str] = []
    try:
        for formula in formulas:
            match = re.fullmatch(
                r"(?:'((?:''|[^'])+)'|([^!]+))!(\$?[A-Z]+\$?\d+(?::\$?[A-Z]+\$?\d+)?)",
                formula.strip(),
            )
            if not match:
                continue
            sheet_name = (match.group(1) or match.group(2) or "").replace("''", "'")
            if sheet_name not in workbook.sheetnames:
                continue
            coordinate = match.group(3).replace("$", "")
            worksheet = workbook[sheet_name]
            cells = worksheet[coordinate]
            if isinstance(cells, tuple):
                flattened = [cell for row in cells for cell in (row if isinstance(row, tuple) else (row,))]
            else:
                flattened = [cells]
            values.extend(str(cell.value).strip() for cell in flattened if cell.value is not None)
    finally:
        workbook.close()
    return values


def _xml_value(node: ET.Element | None) -> str:
    return str(node.attrib.get("val") or "").strip() if node is not None else ""


def _xml_bool(node: ET.Element | None) -> bool:
    return _xml_value(node).casefold() in {"1", "true"}


def _xml_float(node: ET.Element | None) -> float | None:
    value = _xml_value(node)
    if not value:
        return None
    try:
        return float(value)
    except ValueError:
        return None


def _axis_number_format(axis: ET.Element) -> str:
    node = axis.find("c:numFmt", CHART_NS)
    return str(node.attrib.get("formatCode") or "") if node is not None else ""


def _chart_series_color(series: ET.Element) -> str:
    for path in (".//a:solidFill/a:srgbClr", ".//a:ln/a:solidFill/a:srgbClr"):
        node = series.find(path, CHART_NS)
        color = _hex_rgb(node.attrib.get("val")) if node is not None else None
        if color:
            return f"#{color}"
    return "unresolved"


def _chart_audit(part: str, xml_bytes: bytes) -> tuple[dict[str, Any], list[str]]:
    try:
        root = ET.fromstring(xml_bytes)
    except ET.ParseError as exc:
        return {"part": part, "title": "", "parse_valid": False}, [f"{part} is not valid chart XML: {exc}"]

    title = "".join(node.text or "" for node in root.findall(".//c:title//a:t", CHART_NS)).strip()
    if not title:
        title = "".join(node.text or "" for node in root.findall(".//c:title//c:v", CHART_NS)).strip()
    if not title:
        title = "".join(node.text or "" for node in root.findall(".//c:title//c:f", CHART_NS)).strip()
    overlay = _xml_bool(root.find(".//c:title/c:overlay", CHART_NS))
    errors: list[str] = []
    if not title:
        errors.append(f"{part} has no visible chart title")
    if overlay:
        errors.append(f"{part} `{title or 'untitled'}` overlays its title on the plot area")

    label_values, label_formulas = _chart_label_values(root)
    series_colors = [_chart_series_color(series) for series in root.findall(".//c:ser", CHART_NS)]
    resolved_series_colors = {color for color in series_colors if color != "unresolved"}
    color_review_recommended = len(series_colors) >= 2 and len(resolved_series_colors) <= 1

    value_axes: dict[str, ET.Element] = {}
    for axis in root.findall(".//c:valAx", CHART_NS):
        axis_id = _xml_value(axis.find("c:axId", CHART_NS))
        if axis_id:
            value_axes[axis_id] = axis

    composition_audits: list[dict[str, Any]] = []
    for chart_index, bar_chart in enumerate(root.findall(".//c:barChart", CHART_NS), start=1):
        grouping = _xml_value(bar_chart.find("c:grouping", CHART_NS))
        series_count = len(bar_chart.findall("c:ser", CHART_NS))
        axis_ids = {
            _xml_value(node)
            for node in bar_chart.findall("c:axId", CHART_NS)
            if _xml_value(node)
        }
        axes = [axis for axis_id, axis in value_axes.items() if axis_id in axis_ids]
        percent_axis = any("%" in _axis_number_format(axis) for axis in axes)
        normalized_composition = series_count >= 2 and (
            grouping == "percentStacked" or (grouping == "stacked" and percent_axis)
        )
        if not normalized_composition:
            continue

        axis_audits: list[dict[str, Any]] = []
        for axis in axes:
            minimum = _xml_float(axis.find("c:scaling/c:min", CHART_NS))
            maximum = _xml_float(axis.find("c:scaling/c:max", CHART_NS))
            number_format = _axis_number_format(axis)
            axis_audits.append({"min": minimum, "max": maximum, "number_format": number_format})
        has_fixed_percent_axis = any(
            axis["min"] is not None
            and axis["max"] is not None
            and abs(float(axis["min"])) <= 1e-9
            and abs(float(axis["max"]) - 1.0) <= 1e-9
            and "%" in str(axis["number_format"])
            for axis in axis_audits
        )
        if not has_fixed_percent_axis:
            errors.append(
                f"{part} `{title or 'untitled'}` bar chart {chart_index} is a normalized composition but lacks an explicit 0..1 percentage value axis"
            )
        composition_audits.append(
            {
                "chart_index": chart_index,
                "grouping": grouping,
                "series_count": series_count,
                "axes": axis_audits,
                "fixed_0_100_percent_axis": has_fixed_percent_axis,
            }
        )

    return (
        {
            "part": part,
            "title": title,
            "parse_valid": True,
            "title_overlay": overlay,
            "normalized_compositions": composition_audits,
            "series_colors": series_colors,
            "distinct_resolved_series_color_count": len(resolved_series_colors),
            "color_review_recommended": color_review_recommended,
            "plotted_percentile_points": _percentile_points(label_values),
            "_label_formulas": label_formulas,
        },
        errors,
    )


def _project_relative(root: Path, path: Path) -> str:
    try:
        return path.resolve().relative_to(root.resolve()).as_posix()
    except ValueError as exc:
        raise ValueError(f"Managed SQL and outputs must stay inside project root `{root}`: {path}") from exc


def _resolve_project_file(root: Path, value: str | Path) -> Path:
    path = Path(value)
    if not path.is_absolute():
        path = root / path
    path = path.resolve()
    _project_relative(root, path)
    if not path.is_file():
        raise ValueError(f"File not found: {path}")
    return path


def _resolve_project_output(root: Path, value: str | Path) -> Path:
    path = Path(value)
    if not path.is_absolute():
        path = root / path
    path = path.resolve()
    _project_relative(root, path)
    if path.suffix.lower() != ".xlsx":
        raise ValueError("Visualization value refresh output must be an .xlsx file.")
    if path.exists():
        raise ValueError(f"Visualization value refresh never overwrites an existing file: {path}")
    return path


def _read_result_matrix(path: Path) -> tuple[list[str], list[list[Any]]]:
    if path.suffix.lower() == ".csv":
        last_error: Exception | None = None
        for encoding in ("utf-8-sig", "utf-8", "gb18030"):
            try:
                with path.open("r", encoding=encoding, newline="") as handle:
                    rows = list(csv.reader(handle))
                break
            except UnicodeDecodeError as exc:
                last_error = exc
        else:
            if last_error:
                raise last_error
            rows = []
    elif path.suffix.lower() == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - runtime dependency guard
            raise ValueError("XLSX value refresh requires openpyxl.") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            rows = [list(values) for values in workbook.active.iter_rows(values_only=True)]
        finally:
            workbook.close()
    else:
        raise ValueError(f"Unsupported result file type for value refresh: {path.suffix}")

    rows = [row for row in rows if any(str(value or "").strip() for value in row)]
    if not rows:
        raise ValueError(f"Result file has no readable rows: {path}")
    headers = [str(value or "").strip() for value in rows[0]]
    if not headers or any(not value for value in headers):
        raise ValueError("Result value refresh requires a complete, non-empty header row.")
    if len(set(headers)) != len(headers):
        raise ValueError("Result value refresh requires unique column names.")
    data = [(row + [None] * len(headers))[: len(headers)] for row in rows[1:]]
    return headers, data


_REFRESH_TEXT_FIELD_RE = re.compile(
    r"(?:openid|roleid|id$|device|session|account|账号|设备|会话|日期|时间|date|time|编号|代码|code|key)",
    flags=re.I,
)
_REFRESH_NUMERIC_FIELD_RE = re.compile(
    r"(?:base|人数|数量|次数|条数|总数|合计|金额|时长|均值|平均|中位|分位|占比|比例|比率|率|count|amount|duration|average|mean|median|percent|ratio|rate|^p\d+)",
    flags=re.I,
)


def _numeric_value(value: Any) -> float | int | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return value if math.isfinite(float(value)) else None
    text = str(value or "").strip().replace(",", "")
    if not text:
        return None
    percent = text.endswith("%")
    if percent:
        text = text[:-1].strip()
    if re.fullmatch(r"[+-]?0\d+", text):
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    if not math.isfinite(number):
        return None
    number = number / 100 if percent else number
    return int(number) if float(number).is_integer() and not percent else number


def _comparison_value(value: Any, header: str) -> tuple[str, Any]:
    if value is None or str(value).strip() == "":
        return ("empty", "")
    if isinstance(value, (datetime, date)):
        return ("text", value.isoformat())
    if not _REFRESH_TEXT_FIELD_RE.search(header):
        number = _numeric_value(value)
        if number is not None:
            return ("number", round(float(number), 12))
    return ("text", str(value).strip())


def _refresh_value(value: Any, *, header: str, existing_value: Any, number_format: str) -> Any:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, (datetime, date, int, float)) and not isinstance(value, bool):
        return value
    if _REFRESH_TEXT_FIELD_RE.search(header):
        return str(value)
    number = _numeric_value(value)
    if number is None:
        return value
    percent_value = isinstance(value, str) and value.strip().endswith("%")
    if percent_value or isinstance(existing_value, (int, float)) or _REFRESH_NUMERIC_FIELD_RE.search(header):
        return number
    if "%" in str(number_format or ""):
        return number
    return value


def _managed_visualization_lineage(root: Path, visual_path: Path) -> dict[str, Any]:
    rel_visual = _project_relative(root, visual_path)
    index = load_index(root)
    for entry in index.get("entries", []):
        for version in entry.get("versions", []):
            outputs = version.get("derived_outputs", [])
            visual = next(
                (
                    item
                    for item in outputs
                    if isinstance(item, dict)
                    and item.get("kind") == "visualization"
                    and str(item.get("path") or "").replace("\\", "/") == rel_visual
                ),
                None,
            )
            if not visual:
                continue
            result_id = str(visual.get("source_result_id") or "")
            result = next(
                (
                    item
                    for item in outputs
                    if isinstance(item, dict)
                    and item.get("kind") == "result_evidence"
                    and str(item.get("attachment_id") or "") == result_id
                ),
                None,
            )
            if not result:
                raise ValueError("Managed workspace visualization has no exact source result.")
            return {
                "asset_kind": "temporary_query",
                "source_sql_path": str(version.get("path") or ""),
                "source_sql_fingerprint": str(version.get("sql_fingerprint") or ""),
                "visualization_path": rel_visual,
                "visualization_id": str(visual.get("attachment_id") or ""),
                "result_path": str(result.get("path") or ""),
                "result_id": result_id,
            }

    manifest = read_json(manifest_path(root), {})
    for run in manifest.get("run_evidence", []):
        visual = next(
            (
                item
                for item in run.get("derived_outputs", [])
                if isinstance(item, dict)
                and item.get("kind") == "visualization"
                and str(item.get("path") or "").replace("\\", "/") == rel_visual
            ),
            None,
        )
        if not visual:
            continue
        result_id = str(run.get("result_binding_id") or run.get("run_id") or "")
        if str(visual.get("source_result_id") or "") != result_id:
            raise ValueError("Managed formal visualization has no exact source result.")
        return {
            "asset_kind": str(run.get("sql_asset_kind") or "query"),
            "source_sql_path": str(run.get("sql_path") or run.get("source_artifact") or ""),
            "source_sql_fingerprint": str(run.get("source_sql_fingerprint") or ""),
            "visualization_path": rel_visual,
            "visualization_id": str(visual.get("attachment_id") or ""),
            "result_path": str(run.get("evidence_file") or ""),
            "result_id": result_id,
        }
    raise ValueError("Base visualization must be a managed visualization with exact result lineage.")


def _source_table_match(
    workbook: Any,
    *,
    headers: list[str],
    rows: list[list[Any]],
    source_sheet: str = "",
) -> dict[str, Any] | None:
    if source_sheet and source_sheet not in workbook.sheetnames:
        raise ValueError(f"Source sheet not found in base visualization: {source_sheet}")
    worksheets = [workbook[source_sheet]] if source_sheet else list(workbook.worksheets)
    matches: list[dict[str, Any]] = []
    first_header = headers[0]
    for sheet in worksheets:
        max_row = min(sheet.max_row, 100)
        max_col = min(sheet.max_column, 100)
        for row_index in range(1, max_row + 1):
            for col_index in range(1, max_col + 1):
                if str(sheet.cell(row_index, col_index).value or "").strip() != first_header:
                    continue
                observed_headers = [
                    str(sheet.cell(row_index, col_index + offset).value or "").strip()
                    for offset in range(len(headers))
                ]
                if observed_headers != headers:
                    continue
                compatible = True
                formula_cells: list[str] = []
                for row_offset, source_row in enumerate(rows, start=1):
                    for col_offset, source_value in enumerate(source_row):
                        cell = sheet.cell(row_index + row_offset, col_index + col_offset)
                        if cell.data_type == "f":
                            formula_cells.append(cell.coordinate)
                            compatible = False
                            continue
                        if _comparison_value(cell.value, headers[col_offset]) != _comparison_value(
                            source_value, headers[col_offset]
                        ):
                            compatible = False
                if compatible:
                    matches.append(
                        {
                            "sheet_name": sheet.title,
                            "header_row": row_index,
                            "start_column": col_index,
                            "formula_cells": formula_cells,
                        }
                    )
    if not matches:
        return None
    preferred = [item for item in matches if re.search(r"源|原始|raw|data|result", item["sheet_name"], re.I)]
    candidates = preferred or matches
    if len(candidates) != 1:
        raise ValueError(
            "Base visualization contains multiple source-table matches; pass --source-sheet to select one explicitly."
        )
    return candidates[0]


def prepare_visualization_value_refresh(
    *,
    root: Path,
    sql_path: str | Path,
    base_visualization: str | Path,
    result_file: str | Path,
    output_file: str | Path | None = None,
    source_sheet: str = "",
    dry_run: bool = False,
) -> dict[str, Any]:
    root = root.resolve()
    base_workbook = _resolve_project_file(root, base_visualization)
    if base_workbook.suffix.lower() != ".xlsx":
        raise ValueError("Base visualization must be a managed .xlsx workbook.")
    lineage = _managed_visualization_lineage(root, base_workbook)
    old_result = _resolve_project_file(root, lineage["result_path"])
    old_sql = _resolve_project_file(root, lineage["source_sql_path"])
    new_source = resolve_sql_source(root, sql_path)
    new_sql = _resolve_project_file(root, new_source["path"])
    new_result = Path(result_file).resolve()
    if not new_result.is_file() or new_result.suffix.lower() not in RESULT_EXTENSIONS:
        raise ValueError("New SQL result must be an existing .csv or .xlsx file.")

    old_headers, old_rows = _read_result_matrix(old_result)
    new_headers, new_rows = _read_result_matrix(new_result)
    reasons: list[str] = []
    old_logic_fingerprint = logic_fingerprint(old_sql.read_text(encoding="utf-8-sig"))
    new_logic_fingerprint = logic_fingerprint(new_sql.read_text(encoding="utf-8-sig"))
    if old_logic_fingerprint != new_logic_fingerprint:
        reasons.append("sql_logic_changed")
    if old_headers != new_headers:
        reasons.append("result_columns_changed")
    if len(old_rows) != len(new_rows):
        reasons.append("result_row_count_changed")
    cell_count = len(new_headers) * len(new_rows)
    if cell_count > MAX_VALUE_REFRESH_CELLS:
        reasons.append("refresh_cell_budget_exceeded")

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - runtime dependency guard
        raise ValueError("Visualization value refresh requires openpyxl.") from exc
    workbook = load_workbook(base_workbook, data_only=False)
    try:
        source = None
        if not reasons:
            source = _source_table_match(
                workbook,
                headers=old_headers,
                rows=old_rows,
                source_sheet=source_sheet,
            )
            if not source:
                reasons.append("base_source_table_not_found")
        chart_count = sum(len(sheet._charts) for sheet in workbook.worksheets)
        formula_count = sum(
            1
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if cell.data_type == "f"
        )
        plan = {
            "schema_version": VALUE_REFRESH_VERSION,
            "status": "rebuild_required" if reasons else ("eligible" if dry_run else "candidate_ready"),
            "mode": "value_only",
            "base_lineage": lineage,
            "target_sql": {
                "path": str(new_source["path"]),
                "sql_fingerprint": str(new_source["sql_fingerprint"]),
                "logic_fingerprint": new_logic_fingerprint,
            },
            "compatibility": {
                "sql_logic_unchanged": old_logic_fingerprint == new_logic_fingerprint,
                "base_logic_fingerprint": old_logic_fingerprint,
                "columns_unchanged": old_headers == new_headers,
                "row_count_unchanged": len(old_rows) == len(new_rows),
                "source_table_matched": source is not None,
                "cell_count": cell_count,
                "cell_budget": MAX_VALUE_REFRESH_CELLS,
                "rebuild_reasons": reasons,
            },
            "source_table": source or {},
            "preserved": {
                "chart_count": chart_count,
                "formula_count": formula_count,
                "styles_and_conditional_formats": True,
            },
        }
        if reasons or dry_run:
            return plan
        if output_file is None:
            raise ValueError("Non-dry-run value refresh requires --out inside the project root.")
        output = _resolve_project_output(root, output_file)
        output.parent.mkdir(parents=True, exist_ok=True)
        sheet = workbook[source["sheet_name"]]
        for row_offset, source_row in enumerate(new_rows, start=1):
            for col_offset, value in enumerate(source_row):
                cell = sheet.cell(source["header_row"] + row_offset, source["start_column"] + col_offset)
                cell.value = _refresh_value(
                    value,
                    header=new_headers[col_offset],
                    existing_value=cell.value,
                    number_format=cell.number_format,
                )
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
        workbook.save(output)
    finally:
        workbook.close()

    workbook_check = inspect_visual_workbook(
        output,
        required_percentile_fields=list(new_headers),
        result_inspections=[inspect_result_file(new_result)],
    )
    plan["output"] = {
        "path": _project_relative(root, output),
        "absolute_path": str(output),
        "sha256": file_sha256(output),
    }
    plan["preserved"]["chart_count_after"] = workbook_check["chart_count"]
    plan["preserved"]["formula_recalculation_requested"] = True
    plan["workbook_check"] = workbook_check
    return plan


def _formal_artifact(root: Path, rel_sql: str) -> dict[str, Any] | None:
    if (root / "formal_assets" / "index.json").is_file():
        for package_entry in list_formal_asset_packages(root):
            package_id = str(package_entry.get("package_id") or "")
            package = load_formal_asset_package(root, package_id)
            members = [item for item in package.get("members", []) if isinstance(item, dict)]
            member = next(
                (
                    item
                    for item in members
                    if str(item.get("path") or "").replace("\\", "/") == rel_sql
                ),
                None,
            )
            if member is None:
                continue
            role = str(member.get("role") or "").lower()
            if role not in FORMAL_QUERY_ROLES | FORMAL_DASHBOARD_ROLES:
                return None
            sql_path = Path(rel_sql)
            spec_path = sql_path.with_name(f"{sql_path.stem}.spec.json").as_posix()
            meta_path = sql_path.with_name(f"{sql_path.stem}.meta.json").as_posix()
            spec = read_json(root / spec_path, {})
            meta = read_json(root / meta_path, {})
            version_match = re.search(r"v(\d+)\.sql$", sql_path.name, flags=re.I)
            return {
                "kind": "QUERY" if role in FORMAL_QUERY_ROLES else "DASHBOARD",
                "slug": str((spec.get("spec_meta") or {}).get("artifact_slug") or sql_path.parent.name),
                "version": int(version_match.group(1)) if version_match else 0,
                "title": str(meta.get("title") or package.get("title") or sql_path.stem),
                "artifact_state": str(member.get("lifecycle_state") or ""),
                "path": rel_sql,
                "spec_path": spec_path if (root / spec_path).is_file() else "",
                "package_id": package_id,
                "package_directory": str(package.get("directory") or ""),
                "package_member_id": str(member.get("member_id") or ""),
                "package_manifest": package,
                "execution_route": copy.deepcopy(
                    meta.get("execution_route") or spec.get("execution_route") or {}
                ),
            }

    manifest = read_json(manifest_path(root), {})
    for artifact in manifest.get("artifacts", []):
        if not isinstance(artifact, dict):
            continue
        if str(artifact.get("path") or "").replace("\\", "/") == rel_sql:
            return artifact
    return None


def resolve_sql_source(root: Path, sql_value: str | Path) -> dict[str, Any]:
    sql_path = _resolve_project_file(root, sql_value)
    rel_sql = _project_relative(root, sql_path)
    fingerprint = execution_fingerprint(sql_path.read_text(encoding="utf-8-sig"))
    workspace = find_query_reference(root, sql_path, match_fingerprint=True)
    if workspace:
        return {
            "contract_version": SOURCE_VERSION,
            "asset_kind": "temporary_query",
            "path": rel_sql,
            "sql_fingerprint": fingerprint,
            "query_id": str(workspace.get("query_id") or ""),
            "version": int(workspace.get("version") or 0),
            "title": str(workspace.get("title") or ""),
            "summary_plan": copy.deepcopy(workspace.get("summary_plan") or {}),
            "analysis_role": str(workspace.get("analysis_role") or ""),
            "analysis_bundle": copy.deepcopy(workspace.get("analysis_bundle") or {}),
            "execution_route": copy.deepcopy(workspace.get("execution_route") or {}),
        }

    artifact = _formal_artifact(root, rel_sql)
    if not artifact:
        raise ValueError(
            "SQL result binding requires an exact indexed Workspace SQL or Formal Asset Package QUERY/DASHBOARD member."
        )
    kind = str(artifact.get("kind") or "").upper()
    if kind not in {"QUERY", "DASHBOARD"}:
        raise ValueError(f"Result visualization supports QUERY or DASHBOARD SQL, not {kind or 'unknown'}.")
    return {
        "contract_version": SOURCE_VERSION,
        "asset_kind": "query" if kind == "QUERY" else "dashboard",
        "path": rel_sql,
        "sql_fingerprint": fingerprint,
        "slug": str(artifact.get("slug") or sql_path.parent.name),
        "version": int(artifact.get("version") or 0),
        "title": str(artifact.get("title") or ""),
        "artifact_state": str(artifact.get("artifact_state") or artifact.get("status") or ""),
        "execution_route": copy.deepcopy(artifact.get("execution_route") or {}),
        "package_id": str(artifact.get("package_id") or ""),
        "package_directory": str(artifact.get("package_directory") or ""),
        "package_member_id": str(artifact.get("package_member_id") or ""),
        "package_manifest": copy.deepcopy(artifact.get("package_manifest") or {}),
    }


def inspect_visual_workbook(
    path: Path,
    *,
    required_percentile_fields: list[str] | None = None,
    result_inspections: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    if path.suffix.lower() != ".xlsx":
        raise ValueError("Result visualization must be a reusable .xlsx workbook.")
    if not zipfile.is_zipfile(path):
        raise ValueError(f"Visualization workbook is not a valid XLSX ZIP package: {path}")
    with zipfile.ZipFile(path) as workbook:
        names = set(workbook.namelist())
        workbook_xml = workbook.read("xl/workbook.xml").decode("utf-8", errors="replace") if "xl/workbook.xml" in names else ""
        worksheet_names = [
            html.unescape(value)
            for value in re.findall(r'<sheet[^>]*\sname="([^"]+)"', workbook_xml)
        ]
        chart_parts = xlsx_chart_parts(names)
        chart_audits: list[dict[str, Any]] = []
        chart_errors: list[str] = []
        for chart_part in chart_parts:
            audit, errors = _chart_audit(chart_part, workbook.read(chart_part))
            chart_audits.append(audit)
            chart_errors.extend(errors)
    required_fields = list(
        dict.fromkeys(
            str(item).strip()
            for item in (required_percentile_fields or [])
            if _percentile_points([str(item)])
        )
    )
    required_points = _percentile_points(required_fields)
    for audit in chart_audits:
        label_formulas = audit.pop("_label_formulas", [])
        resolved_labels = _resolve_chart_label_formulas(path, label_formulas) if required_points else []
        audit["plotted_percentile_points"] = _percentile_points(
            list(audit.get("plotted_percentile_points") or []) + resolved_labels
        )
    plotted_points = _percentile_points(
        [
            point
            for audit in chart_audits
            for point in audit.get("plotted_percentile_points", [])
        ]
    )
    missing_points = [point for point in required_points if point not in plotted_points]
    percentile_coverage = {
        "available_fields": required_fields,
        "required_points": required_points,
        "plotted_points": plotted_points,
        "missing_points": missing_points,
        "complete": not missing_points,
    }
    merge_audit = _merge_audit(path, result_inspections=list(result_inspections or []))
    single_line_context = _single_line_context_audit(path)
    conditional_format_contrast = _conditional_format_contrast_audit(path)
    if "xl/workbook.xml" not in names:
        raise ValueError("Visualization workbook is missing xl/workbook.xml.")
    worksheets = sorted(name for name in names if re.fullmatch(r"xl/worksheets/sheet\d+\.xml", name))
    if not worksheets:
        raise ValueError("Visualization workbook has no worksheets.")
    if not chart_parts:
        raise ValueError(
            "Visualization workbook has no Excel chart. Build and visually verify at least one decision-useful chart before binding."
        )
    if merge_audit["blocked_merges"]:
        blocked_merges = merge_audit["blocked_merges"]
        raise ValueError(
            "[VIS-LAYOUT-001] Visualization workbook contains unsupported merged cells: "
            + "; ".join(
                f"{item['sheet']}!{item['range']} ({item['reason']})"
                for item in blocked_merges[:8]
            )
            + ". Only parent-group labels and nested header groups inside chart-containing presentation tables may merge; "
            "keep titles, evidence bands, source sheets, metrics, and summary rows unmerged."
        )
    if single_line_context["violations"]:
        violation = single_line_context["violations"][0]
        raise ValueError(
            f"[VIS-LAYOUT-002] {violation['sheet']} row {violation['row']} wraps presentation context in "
            + ", ".join(violation["wrapped_cells"])
            + ". Metadata, context, and evidence rows must stay single-line with wrap_text=false; "
            "reserve empty spill cells, widen the value column, or move the next key-value pair to another row."
        )
    if conditional_format_contrast["violations"]:
        violation = conditional_format_contrast["violations"][0]
        if violation["status"] == "unresolved":
            detail = "contains an endpoint color that cannot be resolved to RGB"
        else:
            detail = (
                f"uses {violation.get('worst_foreground', 'unknown foreground')} text on "
                f"{violation.get('worst_background', 'unknown background')} at "
                f"{float(violation.get('minimum_observed_contrast') or 0):.2f}:1"
            )
        raise ValueError(
            "[VIS-CONTRAST-001] Conditional color scale "
            f"{violation['sheet']}!{violation['range']} {detail}; "
            f"minimum is {conditional_format_contrast['minimum_contrast_ratio']:.2f}:1. "
            "Use one stable dark foreground with lighter scale endpoints; Excel color scales do not reliably switch font color per value."
        )
    if chart_errors:
        raise ValueError("Visualization workbook chart QA failed: " + "; ".join(chart_errors[:8]) + ".")
    if missing_points:
        raise ValueError(
            "Visualization workbook omits available percentile points: "
            + ", ".join(missing_points)
            + ". Every percentile present in the result must be plotted; titles, notes, and source sheets do not count."
        )
    base_coverage = _base_coverage(
        path,
        result_inspections=list(result_inspections or []),
        chart_audits=chart_audits,
    )
    return {
        "format": "xlsx",
        "worksheet_count": len(worksheets),
        "chart_count": len(chart_parts),
        "worksheet_names": worksheet_names,
        "package_valid": True,
        "merged_cell_count": len(merge_audit["allowed_merges"]),
        "allowed_merged_cell_count": len(merge_audit["allowed_merges"]),
        "blocked_merged_cell_count": 0,
        "merge_audit": merge_audit,
        "single_line_context": single_line_context,
        "conditional_format_contrast": conditional_format_contrast,
        "chart_qa_passed": True,
        "chart_audits": chart_audits,
        "percentile_coverage": percentile_coverage,
        "base_coverage": base_coverage,
    }


def _absolute_asset(root: Path, row: dict[str, Any]) -> dict[str, Any]:
    path = str(row.get("path") or "")
    absolute = (root / path).resolve() if path else Path()
    return {
        **row,
        "status": str(row.get("status") or "attached"),
        "absolute_path": str(absolute) if path else "",
        "exists": bool(path and absolute.is_file()),
    }


def _workspace_bind(
    *,
    root: Path,
    source: dict[str, Any],
    result_file: Path,
    workbook: Path,
    result_title: str,
    result_purpose: str,
    visualization_title: str,
    visualization_purpose: str,
    user_confirmed: bool,
    result_inspection: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, Any]]:
    result = attach_derived_output(
        root=root,
        file_path=result_file,
        title=result_title,
        purpose=result_purpose,
        kind="result_evidence",
        source_kind="user_result",
        query_id=str(source["query_id"]),
        version_number=int(source["version"]),
        result_inspection=result_inspection,
    )
    visual = attach_derived_output(
        root=root,
        file_path=workbook,
        title=visualization_title,
        purpose=visualization_purpose,
        kind="visualization",
        source_kind="skill_generated",
        query_id=str(source["query_id"]),
        version_number=int(source["version"]),
        source_result_id=str(result.get("attachment_id") or ""),
    )
    if user_confirmed:
        reference = find_query_reference(root, root / source["path"], match_fingerprint=True) or {}
        if str(reference.get("status") or "") != "promoted":
            transition_query(
                root=root,
                query_id=str(source["query_id"]),
                sql_path=str(source["path"]),
                status="result_confirmed",
                reason="User confirmed the returned result while requesting result visualization.",
                result_status="user_confirmed",
            )
    return result, visual


def _unique_run_id(root: Path, source: dict[str, Any], result_sha: str) -> str:
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    slug = re.sub(r"[^a-z0-9]+", "-", str(source.get("slug") or Path(source["path"]).parent.name).lower()).strip("-")
    base = f"{stamp}_{slug or 'sql-result'}_{result_sha[:8]}"
    run_id = base
    number = 2
    package_members = (source.get("package_manifest") or {}).get("members", [])
    occupied_paths = {
        str(item.get("path") or "")
        for item in package_members
        if isinstance(item, dict)
    }
    while any(f"/{run_id}/" in path for path in occupied_paths):
        run_id = f"{base}-{number}"
        number += 1
    return run_id


def _formal_visual_row(
    *,
    source: dict[str, Any],
    workbook: Path,
    run_id: str,
    rel_path: str,
    title: str,
    purpose: str,
) -> dict[str, Any]:
    retention = full_reusable_output_retention(workbook, "visualization")
    digest = str(retention["source_sha256"])
    return {
        "attachment_id": f"qwo-{digest[:12]}",
        "kind": "visualization",
        "source_kind": "skill_generated",
        "title": title,
        "purpose": purpose,
        "path": rel_path,
        "original_file_name": workbook.name,
        "media_type": mimetypes.guess_type(workbook.name)[0]
        or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "sha256": digest,
        "source_sha256": digest,
        "retention": retention,
        "source_sql_fingerprint": str(source["sql_fingerprint"]),
        "source_result_id": run_id,
        "lineage_status": "exact_result",
        "workbook_manifest": build_workbook_manifest(workbook),
        "preview_status": "not_available",
        "related_queries": [],
        "generation_provenance": build_generation_provenance(
            generator_script="sql_result_visualization.py",
            workflow="formal_result_visualization_binding",
            artifact_kind="SQL_RESULT_VISUALIZATION",
            generated_at=now_iso(),
            source="skill_generated",
            extra={"result_binding_id": run_id, "source_sql_path": source["path"]},
        ),
        "created_at": now_iso(),
    }


def _run_markdown(record: dict[str, Any]) -> str:
    lines = [
        f"# {record['title']}",
        "",
        f"- result_binding_id: {record['result_binding_id']}",
        f"- source_artifact: {record['source_artifact']}",
        f"- sql_asset_kind: {record['sql_asset_kind']}",
        f"- source_sql_fingerprint: {record['source_sql_fingerprint']}",
        f"- status: {record['status']}",
        f"- user_confirmed: {str(record['user_confirmed']).lower()}",
        f"- evidence_file: {record['evidence_file']}",
        f"- result_time_coverage: {json.dumps(record.get('result_time_coverage') or {}, ensure_ascii=False)}",
        f"- visualization_files: {', '.join(item['path'] for item in record['derived_outputs'])}",
        f"- created_at: {record['created_at']}",
        "",
        "## Result Summary",
        "",
        record["result_summary"],
        "",
    ]
    return "\n".join(lines)


def _formal_bind(
    *,
    root: Path,
    source: dict[str, Any],
    result_file: Path,
    workbook: Path,
    result_title: str,
    result_purpose: str,
    visualization_title: str,
    visualization_purpose: str,
    user_confirmed: bool,
    result_inspection: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, Any]]:
    package_id = str(source.get("package_id") or "")
    source_member_id = str(source.get("package_member_id") or "")
    if not package_id or not source_member_id:
        raise ValueError("Formal result binding requires a resolved Formal Asset Package member.")
    package = load_formal_asset_package(root, package_id)
    members = [item for item in package.get("members", []) if isinstance(item, dict)]
    retained = prepare_result_evidence(result_file)
    source_sha = str(retained.retention["source_sha256"])
    workbook_sha = file_sha256(workbook)

    # Reuse an exact immutable binding receipt when the same SQL, result, and
    # workbook were already registered in this Package.
    for member in members:
        if str(member.get("role") or "").lower() != "run_record":
            continue
        record_path = root / str(member.get("path") or "")
        record = read_json(record_path, {})
        outputs = record.get("derived_outputs") if isinstance(record.get("derived_outputs"), list) else []
        exact_visual = next(
            (
                item
                for item in outputs
                if isinstance(item, dict)
                and str(item.get("source_sha256") or "") == workbook_sha
            ),
            None,
        )
        if (
            str(record.get("source_sql_fingerprint") or "") == source["sql_fingerprint"]
            and str((record.get("result_evidence_retention") or {}).get("source_sha256") or "") == source_sha
            and exact_visual is not None
            and (not user_confirmed or bool(record.get("user_confirmed")))
        ):
            result = {
                "status": "reused",
                "result_binding_id": str(record.get("result_binding_id") or record.get("run_id") or ""),
                "attachment_id": str(record.get("result_member_id") or ""),
                "kind": "result_evidence",
                "path": str(record.get("evidence_file") or ""),
                "retention": record.get("result_evidence_retention") or {},
                "result_time_coverage": copy.deepcopy(record.get("result_time_coverage") or {}),
            }
            return result, {**exact_visual, "status": "reused"}

    package_directory = str(package.get("directory") or "")
    run_id = _unique_run_id(root, {**source, "package_manifest": package}, source_sha)
    evidence_target = f"evidence/{run_id}/result{retained.suffix}"
    visual_target = f"outputs/{run_id}/visualization-{workbook_sha[:8]}.xlsx"
    record_target = f"evidence/{run_id}/run.json"
    evidence_rel = f"{package_directory}/members/{evidence_target}"
    visual_rel = f"{package_directory}/members/{visual_target}"
    record_rel = f"{package_directory}/members/{record_target}"

    existing_result = next(
        (
            item
            for item in members
            if str(item.get("role") or "").lower() == "result_evidence"
            and str(item.get("sha256") or "") == retained.stored_sha256
        ),
        None,
    )
    existing_visual = next(
        (
            item
            for item in members
            if str(item.get("role") or "").lower() == "visualization"
            and str(item.get("sha256") or "") == workbook_sha
        ),
        None,
    )
    if existing_result is not None:
        evidence_rel = str(existing_result.get("path") or "")
    if existing_visual is not None:
        visual_rel = str(existing_visual.get("path") or "")

    def member_id(prefix: str, target: str) -> str:
        return f"{prefix}-{hashlib.sha256(target.encode('utf-8')).hexdigest()[:20]}"

    result_member_id = str((existing_result or {}).get("member_id") or member_id("RESULT", evidence_target))
    visual_member_id = str((existing_visual or {}).get("member_id") or member_id("VISUAL", visual_target))
    record_member_id = member_id("RUN", record_target)
    visual = _formal_visual_row(
        source=source,
        workbook=workbook,
        run_id=result_member_id,
        rel_path=visual_rel,
        title=visualization_title,
        purpose=visualization_purpose,
    )
    created_at = now_iso()
    facts = analyze_sql_file(root / source["path"], kind=str(source["asset_kind"]).upper())
    record = {
        "contract_version": BINDING_VERSION,
        "run_id": run_id,
        "result_binding_id": run_id,
        "package_id": package_id,
        "source_member_id": source_member_id,
        "result_member_id": result_member_id,
        "visualization_member_id": visual_member_id,
        "title": result_title,
        "source_artifact": source["path"],
        "sql_path": source["path"],
        "sql_asset_kind": source["asset_kind"],
        "source_sql_fingerprint": source["sql_fingerprint"],
        "parameter_snapshot": facts.get("params") or {},
        "status": "passed" if user_confirmed else "observed",
        "row_count": retained.retention.get("source_row_count"),
        "checked_metrics": [],
        "checked_dimensions": [],
        "sample_fields": list(retained.retention.get("columns") or []),
        "result_summary": result_purpose,
        "issues": "",
        "user_confirmed": bool(user_confirmed),
        "confirmed_by": "user" if user_confirmed else "",
        "evidence_file": evidence_rel,
        "result_file_type": retained.suffix,
        "result_evidence_retention": retained.retention,
        "result_time_coverage": copy.deepcopy(result_inspection.get("time_coverage") or {}),
        "derived_outputs": [visual],
        "generation_provenance": build_generation_provenance(
            generator_script="sql_result_visualization.py",
            workflow="formal_asset_package_result_binding",
            artifact_kind="SQL_RESULT_EVIDENCE",
            generated_at=created_at,
            source="user_result",
            extra={"result_binding_id": run_id, "source_sql_path": source["path"]},
        ),
        "created_at": created_at,
        "notes": "Returned SQL result and visualization bound to one exact Package SQL member.",
        "path": record_rel,
    }

    temp_parent = root / ".tmp"
    temp_parent.mkdir(exist_ok=True)
    with tempfile.TemporaryDirectory(prefix="formal-result-", dir=temp_parent) as tmp:
        temp_root = Path(tmp)
        new_members: list[dict[str, Any]] = []
        if existing_result is None:
            retained_path = temp_root / f"result{retained.suffix}"
            retained_path.write_bytes(retained.payload)
            new_members.append(
                {
                    "member_id": result_member_id,
                    "source_path": retained_path,
                    "target_path": evidence_target,
                    "role": "result_evidence",
                    "lifecycle_state": "current",
                }
            )
        if existing_visual is None:
            new_members.append(
                {
                    "member_id": visual_member_id,
                    "source_path": workbook,
                    "target_path": visual_target,
                    "role": "visualization",
                    "lifecycle_state": "current",
                }
            )
        record_path = temp_root / "run.json"
        record_path.write_text(json.dumps(record, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        new_members.append(
            {
                "member_id": record_member_id,
                "source_path": record_path,
                "target_path": record_target,
                "role": "run_record",
                "lifecycle_state": "current",
            }
        )
        lineage = [
            {"relation": "evidence_for", "from_member_id": result_member_id, "to_member_id": source_member_id},
            {"relation": "evidenced_by", "from_member_id": source_member_id, "to_member_id": result_member_id},
            {"relation": "derived_from_result", "from_member_id": visual_member_id, "to_member_id": result_member_id},
            {"relation": "has_visualization", "from_member_id": result_member_id, "to_member_id": visual_member_id},
            {"relation": "records_run_for", "from_member_id": record_member_id, "to_member_id": source_member_id},
            {"relation": "records_evidence", "from_member_id": record_member_id, "to_member_id": result_member_id},
            {"relation": "records_output", "from_member_id": record_member_id, "to_member_id": visual_member_id},
        ]
        plan = plan_formal_asset_package(
            root,
            title=str(package.get("title") or source.get("title") or package_id),
            members=new_members,
            package_id=package_id,
            lineage=lineage,
            lifecycle_state=str(package.get("lifecycle_state") or "current"),
        )
        repository_receipt = apply_formal_asset_plan(plan)
        validation = validate_formal_asset_receipt(root, repository_receipt)
        if validation.get("status") != "valid":
            raise ValueError(
                "Formal Asset Repository returned an invalid result binding receipt: "
                + "; ".join(str(item) for item in validation.get("problems", []))
            )

    result = {
        "status": "reused" if existing_result is not None else "attached",
        "result_binding_id": run_id,
        "attachment_id": result_member_id,
        "kind": "result_evidence",
        "path": evidence_rel,
        "retention": retained.retention,
        "result_time_coverage": copy.deepcopy(result_inspection.get("time_coverage") or {}),
        "formal_asset_receipt": repository_receipt,
    }
    return result, visual


def bind_result_visualization(
    *,
    root: Path,
    sql_path: str | Path,
    result_file: str | Path,
    visualization_file: str | Path,
    result_title: str,
    result_purpose: str,
    visualization_title: str,
    visualization_purpose: str,
    user_confirmed: bool,
) -> dict[str, Any]:
    root = root.resolve()
    source = resolve_sql_source(root, sql_path)
    if (source.get("summary_plan") or {}).get("routing") == "grouped_plus_overall":
        bundle_id = str((source.get("analysis_bundle") or {}).get("bundle_id") or "")
        raise ValueError(
            "This SQL is one member of a grouped/overall analysis bundle. "
            f"Use attach-bundle-result and bind-bundle for `{bundle_id or 'the linked bundle'}`."
        )
    result_path = Path(result_file).resolve()
    workbook = Path(visualization_file).resolve()
    if not result_path.is_file() or result_path.suffix.lower() not in RESULT_EXTENSIONS:
        raise ValueError("Returned SQL result must be an existing .csv or .xlsx file.")
    sql_text = (root / source["path"]).read_text(encoding="utf-8-sig")
    project_config = read_json(root / "project_config.json", {})
    effective_config, _ = effective_config_for_context(
        project_config,
        sql_text,
        source.get("execution_route"),
    )
    result_inspection = inspect_result_file(
        result_path,
        sql=sql_text,
        project_config=effective_config,
    )
    time_coverage_blockers = time_coverage_problem_messages(
        result_inspection.get("time_coverage")
    )
    if user_confirmed and time_coverage_blockers:
        raise ValueError(" ".join(time_coverage_blockers))
    workbook_check = inspect_visual_workbook(
        workbook,
        required_percentile_fields=list(result_inspection.get("columns") or []),
        result_inspections=[result_inspection],
    )
    if file_sha256(result_path) == file_sha256(workbook):
        raise ValueError("Visualization workbook must be a distinct generated artifact, not the unchanged returned result file.")

    if source["asset_kind"] == "temporary_query":
        result, visual = _workspace_bind(
            root=root,
            source=source,
            result_file=result_path,
            workbook=workbook,
            result_title=result_title,
            result_purpose=result_purpose,
            visualization_title=visualization_title,
            visualization_purpose=visualization_purpose,
            user_confirmed=user_confirmed,
            result_inspection=result_inspection,
        )
    else:
        result, visual = _formal_bind(
            root=root,
            source=source,
            result_file=result_path,
            workbook=workbook,
            result_title=result_title,
            result_purpose=result_purpose,
            visualization_title=visualization_title,
            visualization_purpose=visualization_purpose,
            user_confirmed=user_confirmed,
            result_inspection=result_inspection,
        )

    result_asset = _absolute_asset(root, result)
    visual_asset = _absolute_asset(root, visual)
    status = "ready" if result_asset["exists"] and visual_asset["exists"] else "blocked"
    public_source = {
        key: value
        for key, value in source.items()
        if key != "package_manifest"
    }
    return {
        "schema_version": RECEIPT_VERSION,
        "status": status,
        "sql_asset": {**public_source, "absolute_path": str((root / source["path"]).resolve())},
        "result_asset": result_asset,
        "visualization_asset": visual_asset,
        "checks": {
            "exact_sql_version_bound": True,
            "result_bound": result_asset["exists"],
            "visualization_bound_to_result": visual.get("source_result_id")
            in {result.get("attachment_id"), result.get("result_binding_id")},
            "workbook": workbook_check,
            "sql_version_changed": False,
            "result_time_coverage": copy.deepcopy(
                result_inspection.get("time_coverage") or {}
            ),
        },
    }


def _bundle_member(bundle: dict[str, Any], role: str) -> dict[str, Any]:
    member = next(
        (
            item
            for item in bundle.get("members", [])
            if isinstance(item, dict) and item.get("role") == role
        ),
        None,
    )
    if not member:
        raise ValueError(f"Analysis bundle has no `{role}` member.")
    return member


def _field_key(value: str) -> str:
    return re.sub(r"[\s`\"'\[\]]+", "", str(value or "")).casefold()


def _workspace_result_row(root: Path, binding: dict[str, Any]) -> dict[str, Any]:
    index = load_index(root)
    query_id = str(binding.get("query_id") or "")
    version_number = int(binding.get("version") or 0)
    result_id = str(binding.get("result_id") or "")
    for entry in index.get("entries", []):
        if not isinstance(entry, dict) or entry.get("query_id") != query_id:
            continue
        for version in entry.get("versions", []):
            if not isinstance(version, dict) or int(version.get("version") or 0) != version_number:
                continue
            output = next(
                (
                    item
                    for item in version.get("derived_outputs", [])
                    if isinstance(item, dict)
                    and item.get("kind") == "result_evidence"
                    and item.get("attachment_id") == result_id
                ),
                None,
            )
            if output:
                return output
    raise ValueError(f"Bundle result binding does not resolve: {query_id}@v{version_number}:{result_id}")


def _bundle_status(bundle: dict[str, Any]) -> str:
    roles = {str(item.get("role") or "") for item in bundle.get("members", []) if isinstance(item, dict)}
    bound = set((bundle.get("result_bindings") or {}).keys())
    if roles and roles <= bound:
        return "ready_for_visualization"
    return "awaiting_results"


def _bundle_output_reference(root: Path, bundle_path: Path, bundle: dict[str, Any]) -> dict[str, Any]:
    return {
        "contract_version": BUNDLE_OUTPUT_REF_VERSION,
        "bundle_id": str(bundle.get("bundle_id") or ""),
        "path": bundle_path.relative_to(root).as_posix(),
        "metric_contract_fingerprint": str(bundle.get("metric_contract_fingerprint") or ""),
    }


def attach_analysis_bundle_result(
    *,
    root: Path,
    bundle_value: str | Path,
    role: str,
    result_file: str | Path,
    user_confirmed: bool,
) -> dict[str, Any]:
    """Bind one returned result to its exact role/SQL in an analysis bundle."""

    root = root.resolve()
    bundle_path, bundle = load_analysis_bundle(root, bundle_value)
    role = str(role or "").strip().lower()
    if role not in {"grouped", "overall"}:
        raise ValueError("Bundle result role must be grouped or overall.")
    member = _bundle_member(bundle, role)
    result_path = Path(result_file).resolve()
    if not result_path.is_file() or result_path.suffix.lower() not in RESULT_EXTENSIONS:
        raise ValueError("Returned SQL result must be an existing .csv or .xlsx file.")
    inspection = inspect_result_file(result_path)
    expected = {_field_key(item) for item in member.get("expected_fields", [])}
    actual = {_field_key(item) for item in inspection.get("columns", [])}
    missing = [item for item in member.get("expected_fields", []) if _field_key(item) not in actual]
    if missing:
        raise ValueError(f"{role} result is missing expected SQL fields: {', '.join(missing)}")
    if not expected:
        raise ValueError(f"{role} SQL member has no expected output fields.")

    retained = prepare_result_evidence(result_path)
    existing_binding = (bundle.get("result_bindings") or {}).get(role)
    if existing_binding:
        if str(existing_binding.get("source_sha256") or "") != str(retained.retention.get("source_sha256") or ""):
            raise ValueError(f"Bundle role `{role}` already has a different immutable result binding.")
        output = _workspace_result_row(root, existing_binding)
        attach_status = "reused"
    else:
        attached = attach_derived_output(
            root=root,
            file_path=result_path,
            title=f"{bundle.get('title', '分析')} - {'分组结果' if role == 'grouped' else '整体结果'}",
            purpose=f"绑定 {role} SQL 返回结果，供同一分析 bundle 的最终可视化使用。",
            kind="result_evidence",
            source_kind="user_result",
            query_id=str(member.get("query_id") or ""),
            version_number=int(member.get("version") or 0),
        )
        binding_stub = {
            "query_id": str(member.get("query_id") or ""),
            "version": int(member.get("version") or 0),
            "result_id": str(attached.get("attachment_id") or ""),
        }
        output = _workspace_result_row(root, binding_stub)
        attach_status = str(attached.get("status") or "attached")
        existing_binding = {
            **binding_stub,
            "path": str(output.get("path") or ""),
            "source_sha256": str(output.get("source_sha256") or output.get("sha256") or ""),
            "columns": list((output.get("retention") or {}).get("columns") or inspection.get("columns") or []),
            "attached_at": now_iso(),
        }
        bundle.setdefault("result_bindings", {})[role] = existing_binding
    if user_confirmed:
        transition_query(
            root=root,
            query_id=str(member.get("query_id") or ""),
            status="result_confirmed",
            reason=f"User confirmed the returned {role} result for analysis bundle {bundle.get('bundle_id')}.",
            result_status="user_confirmed",
        )
    bundle["status"] = "visualized" if bundle.get("visualization") else _bundle_status(bundle)
    bundle["updated_at"] = now_iso()
    _write_transaction({bundle_path: json_text(bundle)})
    missing_roles = [
        item.get("role")
        for item in bundle.get("members", [])
        if item.get("role") not in bundle.get("result_bindings", {})
    ]
    return {
        "schema_version": BUNDLE_RESULT_RECEIPT_VERSION,
        "status": "ready_for_visualization" if not missing_roles else "awaiting_other_result",
        "bundle_id": bundle.get("bundle_id"),
        "bundle_path": bundle_path.relative_to(root).as_posix(),
        "role": role,
        "attach_status": attach_status,
        "result_asset": _absolute_asset(root, output),
        "missing_roles": missing_roles,
        "result_columns": inspection.get("columns", []),
    }


def bind_analysis_bundle_visualization(
    *,
    root: Path,
    bundle_value: str | Path,
    visualization_file: str | Path,
    visualization_title: str,
    visualization_purpose: str,
) -> dict[str, Any]:
    """Bind one verified workbook to every exact result in an analysis bundle."""

    root = root.resolve()
    bundle_path, bundle = load_analysis_bundle(root, bundle_value)
    if _bundle_status(bundle) != "ready_for_visualization":
        missing = [
            item.get("role")
            for item in bundle.get("members", [])
            if item.get("role") not in bundle.get("result_bindings", {})
        ]
        raise ValueError("Analysis bundle is missing exact result bindings: " + ", ".join(str(item) for item in missing))
    workbook = Path(visualization_file).resolve()
    result_bindings = bundle.get("result_bindings", {})
    percentile_fields = [
        str(column)
        for role in ("grouped", "overall")
        for binding in [result_bindings.get(role) or {}]
        for column in binding.get("columns", [])
    ]
    result_inspections = [
        inspect_result_file(root / str(_workspace_result_row(root, binding).get("path") or ""))
        for role in ("grouped", "overall")
        for binding in [result_bindings.get(role) or {}]
    ]
    workbook_check = inspect_visual_workbook(
        workbook,
        required_percentile_fields=percentile_fields,
        result_inspections=result_inspections,
    )
    required_sheets = {"分组结果", "整体结果"}
    actual_sheets = set(workbook_check.get("worksheet_names") or [])
    if not required_sheets <= actual_sheets:
        raise ValueError("Bundle visualization workbook must contain separate `分组结果` and `整体结果` source sheets.")

    grouped = _bundle_member(bundle, "grouped")
    result_references = [
        {
            "query_id": str(binding.get("query_id") or ""),
            "version": int(binding.get("version") or 0),
            "result_id": str(binding.get("result_id") or ""),
        }
        for role in ["grouped", "overall"]
        for binding in [bundle.get("result_bindings", {}).get(role) or {}]
    ]
    visual = attach_derived_output(
        root=root,
        file_path=workbook,
        title=visualization_title,
        purpose=visualization_purpose,
        kind="visualization",
        source_kind="skill_generated",
        query_id=str(grouped.get("query_id") or ""),
        version_number=int(grouped.get("version") or 0),
        source_result_references=result_references,
        analysis_bundle_reference=_bundle_output_reference(root, bundle_path, bundle),
    )
    output = {
        "query_id": str(grouped.get("query_id") or ""),
        "version": int(grouped.get("version") or 0),
        "attachment_id": str(visual.get("attachment_id") or ""),
        "path": str(visual.get("path") or ""),
        "kind": "visualization",
        "source_result_ids": [item["result_id"] for item in result_references],
    }
    bundle["visualization"] = output
    bundle["status"] = "visualized"
    bundle["updated_at"] = now_iso()
    _write_transaction({bundle_path: json_text(bundle)})
    visual_asset = _absolute_asset(root, visual)
    return {
        "schema_version": BUNDLE_RECEIPT_VERSION,
        "status": "ready" if visual_asset.get("exists") else "blocked",
        "bundle_id": bundle.get("bundle_id"),
        "bundle_path": bundle_path.relative_to(root).as_posix(),
        "sql_assets": copy.deepcopy(bundle.get("members") or []),
        "result_assets": [
            _absolute_asset(root, _workspace_result_row(root, binding))
            for binding in bundle.get("result_bindings", {}).values()
        ],
        "visualization_asset": visual_asset,
        "checks": {
            "all_bundle_results_bound": True,
            "lineage_status": "exact_results",
            "source_result_count": len(result_references),
            "required_source_sheets": sorted(required_sheets),
            "workbook": workbook_check,
            "sql_version_changed": False,
        },
    }


def _legacy_output_lineage(outputs: list[dict[str, Any]]) -> tuple[int, int, int]:
    results = [item for item in outputs if item.get("kind") == "result_evidence"]
    exact = unresolved = changed = 0
    for output in outputs:
        if output.get("kind") == "result_evidence":
            source_result_id = str(output.get("attachment_id") or "")
            status = "result_evidence"
        elif len(results) == 1:
            source_result_id = str(results[0].get("attachment_id") or "")
            status = "exact_result"
            exact += 1
        else:
            source_result_id = ""
            status = "unresolved_legacy"
            unresolved += 1
        if output.get("source_result_id") != source_result_id or output.get("lineage_status") != status:
            output["source_result_id"] = source_result_id
            output["lineage_status"] = status
            changed += 1
    return exact, unresolved, changed


def migrate_result_lineage(root: Path, *, dry_run: bool) -> dict[str, Any]:
    root = root.resolve()
    files: dict[Path, str | bytes] = {}
    stats = {
        "workspace_outputs_changed": 0,
        "workspace_exact_visual_lineage": 0,
        "workspace_unresolved_legacy": 0,
        "formal_runs_changed": 0,
        "formal_runs_unresolved": 0,
    }
    index_path = root / INDEX_REL
    if index_path.exists():
        index = load_index(root)
        for entry in index.get("entries", []):
            for version in entry.get("versions", []):
                outputs = version.get("derived_outputs")
                if not isinstance(outputs, list) or not outputs:
                    continue
                exact, unresolved, changed = _legacy_output_lineage(outputs)
                stats["workspace_exact_visual_lineage"] += exact
                stats["workspace_unresolved_legacy"] += unresolved
                stats["workspace_outputs_changed"] += changed
                if changed:
                    meta_path = root / str(version.get("meta_path") or "")
                    meta = read_json(meta_path, {})
                    meta["derived_outputs"] = copy.deepcopy(outputs)
                    files[meta_path] = json.dumps(meta, ensure_ascii=False, indent=2) + "\n"
        if stats["workspace_outputs_changed"]:
            files.update(_index_files(root, index))

    manifest_file = manifest_path(root)
    manifest = read_json(manifest_file, {})
    artifacts = {
        str(item.get("path") or ""): item
        for item in manifest.get("artifacts", [])
        if isinstance(item, dict)
    }
    for run in manifest.get("run_evidence", []):
        if not isinstance(run, dict) or not str(run.get("evidence_file") or ""):
            continue
        source_path = str(run.get("sql_path") or run.get("source_artifact") or "")
        source_file = root / source_path
        if not source_file.is_file():
            stats["formal_runs_unresolved"] += 1
            continue
        artifact = artifacts.get(source_path, {})
        kind = str(artifact.get("kind") or "").upper()
        if kind not in {"QUERY", "DASHBOARD"}:
            stats["formal_runs_unresolved"] += 1
            continue
        expected = {
            "contract_version": BINDING_VERSION,
            "result_binding_id": str(run.get("run_id") or ""),
            "sql_asset_kind": "query" if kind == "QUERY" else "dashboard",
            "source_sql_fingerprint": execution_fingerprint(source_file.read_text(encoding="utf-8-sig")),
        }
        changed = any(run.get(key) != value for key, value in expected.items())
        for key, value in expected.items():
            run[key] = value
        if not isinstance(run.get("derived_outputs"), list):
            run["derived_outputs"] = []
            changed = True
        if changed:
            stats["formal_runs_changed"] += 1
    if stats["formal_runs_changed"]:
        manifest["updated_at"] = now_iso()
        files[manifest_file] = json.dumps(manifest, ensure_ascii=False, indent=2) + "\n"

    if files and not dry_run:
        _write_transaction(files)
        rebuild_index(root)
    return {
        "schema_version": "sql_result_lineage_migration_v1",
        "status": "planned" if dry_run else "migrated",
        "project_root": str(root),
        "changed_file_count": len(files),
        "stats": stats,
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    sub = parser.add_subparsers(dest="command", required=True)

    bind = sub.add_parser("bind", help="Bind one returned result and its verified visual Excel to an exact SQL version")
    bind.add_argument("--root", required=True)
    bind.add_argument("--sql-path", required=True)
    bind.add_argument("--result-file", required=True)
    bind.add_argument("--visualization-file", required=True)
    bind.add_argument("--result-title", required=True)
    bind.add_argument("--result-purpose", required=True)
    bind.add_argument("--visualization-title", required=True)
    bind.add_argument("--visualization-purpose", required=True)
    bind.add_argument("--user-confirmed", action="store_true")
    bind.add_argument("--format", choices=["json", "text"], default="json")
    add_function_gate_arguments(bind, selection_help="Use [RESULT_VISUALIZATION].")

    refresh_values = sub.add_parser(
        "refresh-values",
        help="Decide whether a managed visualization can reuse its layout, then replace only the exact source-table values",
    )
    refresh_values.add_argument("--root", required=True)
    refresh_values.add_argument("--sql-path", required=True, help="Exact indexed SQL version that produced the new result")
    refresh_values.add_argument("--base-visualization", required=True, help="Managed prior visualization with exact result lineage")
    refresh_values.add_argument("--result-file", required=True, help="New returned result with the same value-only contract")
    refresh_values.add_argument("--out", help="Project-local candidate .xlsx path; required unless --dry-run")
    refresh_values.add_argument("--source-sheet", default="", help="Explicit source sheet when more than one exact table match exists")
    refresh_values.add_argument("--dry-run", action="store_true")
    refresh_values.add_argument("--format", choices=["json", "text"], default="json")
    add_function_gate_arguments(refresh_values, selection_help="Use [RESULT_VISUALIZATION].")

    attach_bundle = sub.add_parser(
        "attach-bundle-result",
        help="Bind one grouped or overall result to its exact SQL member before combined visualization",
    )
    attach_bundle.add_argument("--root", required=True)
    attach_bundle.add_argument("--bundle", required=True, help="Bundle id or project-local query_analysis_bundle_v1 path")
    attach_bundle.add_argument("--role", choices=["grouped", "overall"], required=True)
    attach_bundle.add_argument("--result-file", required=True)
    attach_bundle.add_argument("--user-confirmed", action="store_true")
    attach_bundle.add_argument("--format", choices=["json", "text"], default="json")
    add_function_gate_arguments(attach_bundle, selection_help="Use [RESULT_VISUALIZATION].")

    bind_bundle = sub.add_parser(
        "bind-bundle",
        help="Bind one verified workbook to all exact grouped/overall results in a bundle",
    )
    bind_bundle.add_argument("--root", required=True)
    bind_bundle.add_argument("--bundle", required=True, help="Bundle id or project-local query_analysis_bundle_v1 path")
    bind_bundle.add_argument("--visualization-file", required=True)
    bind_bundle.add_argument("--visualization-title", required=True)
    bind_bundle.add_argument("--visualization-purpose", required=True)
    bind_bundle.add_argument("--format", choices=["json", "text"], default="json")
    add_function_gate_arguments(bind_bundle, selection_help="Use [RESULT_VISUALIZATION].")

    migrate = sub.add_parser("migrate", help="Backfill result-to-visual lineage without guessing ambiguous history")
    migrate.add_argument("--root", required=True)
    migrate.add_argument("--dry-run", action="store_true")
    migrate.add_argument("--format", choices=["json", "text"], default="json")
    add_function_gate_arguments(migrate, selection_help="Use [RESULT_VISUALIZATION], [PROJECT_ADMIN], or [SKILL_EVOLUTION].")
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    try:
        require_user_function_selection(
            args.function_selection,
            user_request=args.user_request,
            allowed_ids=command_function_ids("sql_result_visualization.py", args.command),
            purpose=f"sql_result_visualization.py {args.command}",
        )
        require_user_request(args.user_request, purpose=f"sql_result_visualization.py {args.command}")
        if args.command == "bind":
            result = bind_result_visualization(
                root=Path(args.root),
                sql_path=args.sql_path,
                result_file=args.result_file,
                visualization_file=args.visualization_file,
                result_title=args.result_title,
                result_purpose=args.result_purpose,
                visualization_title=args.visualization_title,
                visualization_purpose=args.visualization_purpose,
                user_confirmed=bool(args.user_confirmed),
            )
        elif args.command == "refresh-values":
            result = prepare_visualization_value_refresh(
                root=Path(args.root),
                sql_path=args.sql_path,
                base_visualization=args.base_visualization,
                result_file=args.result_file,
                output_file=args.out,
                source_sheet=args.source_sheet,
                dry_run=bool(args.dry_run),
            )
        elif args.command == "attach-bundle-result":
            result = attach_analysis_bundle_result(
                root=Path(args.root),
                bundle_value=args.bundle,
                role=args.role,
                result_file=args.result_file,
                user_confirmed=bool(args.user_confirmed),
            )
        elif args.command == "bind-bundle":
            result = bind_analysis_bundle_visualization(
                root=Path(args.root),
                bundle_value=args.bundle,
                visualization_file=args.visualization_file,
                visualization_title=args.visualization_title,
                visualization_purpose=args.visualization_purpose,
            )
        else:
            result = migrate_result_lineage(Path(args.root), dry_run=bool(args.dry_run))
        if args.format == "json":
            print(json.dumps(result, ensure_ascii=False, indent=2))
        else:
            print(f"{result.get('status')}: {result.get('schema_version')}")
        return 0 if result.get("status") not in {"blocked", "rebuild_required"} else 1
    except FunctionGateError as exc:
        return exit_with_gate_error(exc)
    except (OSError, ValueError, json.JSONDecodeError) as exc:
        parser.error(str(exc))
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
