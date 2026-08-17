from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font


SCRIPT_DIR = Path(__file__).resolve().parents[1] / "scripts"
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

import validate_style  # noqa: E402


class ValidateStyleTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)
        self.result = self.root / "result.csv"
        self.result.write_text(
            "分组,占比,玩家数\nA,0.6,600\nB,0.4,400\n",
            encoding="utf-8-sig",
        )

    def tearDown(self) -> None:
        self.temp.cleanup()

    def _workbook(self, *, show_base: bool) -> Path:
        path = self.root / ("pass.xlsx" if show_base else "blocked.xlsx")
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "分析展示"
        sheet.append(["分组", "玩家数" if show_base else None, "占比"])
        sheet.append(["A", 600 if show_base else None, 0.6])
        sheet.append(["B", 400 if show_base else None, 0.4])
        chart = BarChart()
        chart.title = "玩家占比 × 分组"
        chart.add_data(Reference(sheet, min_col=3, min_row=1, max_row=3), titles_from_data=True)
        chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=3))
        sheet.add_chart(chart, "E2")
        workbook.save(path)
        return path

    def _workbook_with_percentage_scale(self, *, high_color: str) -> Path:
        path = self.root / f"scale-{high_color}.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "分析展示"
        sheet.append(["分组", "玩家数", "占比"])
        sheet.append(["A", 600, 0.6])
        sheet.append(["B", 400, 0.4])
        for cell in (sheet["C2"], sheet["C3"]):
            cell.number_format = "0.0%"
            cell.font = Font(color="17212B")
        sheet.conditional_formatting.add(
            "C2:C3",
            ColorScaleRule(
                start_type="num",
                start_value=0,
                start_color="FFFFFF",
                mid_type="percentile",
                mid_value=50,
                mid_color="D2DFE9",
                end_type="max",
                end_color=high_color,
            ),
        )
        chart = BarChart()
        chart.title = "玩家占比 × 分组"
        chart.add_data(Reference(sheet, min_col=3, min_row=1, max_row=3), titles_from_data=True)
        chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=3))
        sheet.add_chart(chart, "E2")
        workbook.save(path)
        return path

    def _workbook_with_series_colors(self, colors: list[str]) -> tuple[Path, Path]:
        result = self.root / ("series-" + "-".join(colors) + ".csv")
        result.write_text(
            "分组,玩家数,A占比,B占比\nA,600,0.6,0.4\nB,400,0.4,0.6\n",
            encoding="utf-8-sig",
        )
        path = self.root / ("series-" + "-".join(colors) + ".xlsx")
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "分析展示"
        sheet.append(["分组", "玩家数", "A占比", "B占比"])
        sheet.append(["A", 600, 0.6, 0.4])
        sheet.append(["B", 400, 0.4, 0.6])
        chart = BarChart()
        chart.title = "A占比、B占比 × 分组"
        chart.add_data(Reference(sheet, min_col=3, max_col=4, min_row=1, max_row=3), titles_from_data=True)
        chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=3))
        for series, color in zip(chart.series, colors):
            series.graphicalProperties.solidFill = color
        sheet.add_chart(chart, "F2")
        workbook.save(path)
        return path, result

    def test_prebind_validator_reuses_base_gate(self) -> None:
        blocked = validate_style.validate_workbook(self._workbook(show_base=False), [self.result])
        passed = validate_style.validate_workbook(self._workbook(show_base=True), [self.result])

        self.assertEqual(blocked["status"], "blocked")
        self.assertEqual(blocked["blockers"][0]["rule_id"], "VIS-BASE-002")
        self.assertEqual(passed["status"], "pass")
        self.assertTrue(passed["checks"]["base_coverage"]["complete"])
        self.assertEqual(passed["tokens_version"], "viz_tokens_v3")
        self.assertEqual(validate_style.VIZ_TOKENS["conditional_formatting"]["percentage_color_scale"]["kind"], "color_scale")
        self.assertEqual(validate_style.VIZ_TOKENS["conditional_formatting"]["percentage_color_scale"]["foreground"], "#17212B")
        self.assertEqual(validate_style.VIZ_TOKENS["conditional_formatting"]["percentage_color_scale"]["minimum_contrast_ratio"], 4.5)
        self.assertEqual(validate_style.VIZ_TOKENS["conditional_formatting"]["numeric_data_bar"]["kind"], "data_bar")

    def test_percentage_color_scale_with_readable_endpoint_passes_and_is_audited(self) -> None:
        receipt = validate_style.validate_workbook(
            self._workbook_with_percentage_scale(high_color="6F96B8"),
            [self.result],
        )

        self.assertEqual(receipt["status"], "pass")
        audit = receipt["checks"]["conditional_format_contrast"]
        self.assertTrue(audit["complete"])
        self.assertEqual(audit["color_scale_rule_count"], 1)
        self.assertEqual(audit["audited_range_count"], 1)
        self.assertEqual(audit["violation_count"], 0)
        self.assertEqual(audit["audits"][0]["sheet"], "分析展示")
        self.assertEqual(audit["audits"][0]["range"], "C2:C3")
        self.assertEqual(audit["audits"][0]["semantic_kind"], "percentage_or_rate")
        self.assertGreaterEqual(audit["audits"][0]["minimum_observed_contrast"], 4.5)

    def test_percentage_color_scale_with_dark_endpoint_is_blocked(self) -> None:
        receipt = validate_style.validate_workbook(
            self._workbook_with_percentage_scale(high_color="24445D"),
            [self.result],
        )

        self.assertEqual(receipt["status"], "blocked")
        self.assertEqual(receipt["blockers"][0]["rule_id"], "VIS-CONTRAST-001")
        self.assertIn("分析展示!C2:C3", receipt["blockers"][0]["message"])
        self.assertIn("#17212B", receipt["blockers"][0]["message"])
        self.assertIn("#24445D", receipt["blockers"][0]["message"])

    def test_multi_series_single_color_is_flagged_for_visual_review(self) -> None:
        workbook, result = self._workbook_with_series_colors(["24445D", "24445D"])

        receipt = validate_style.validate_workbook(workbook, [result])

        self.assertEqual(receipt["status"], "pass")
        audit = receipt["checks"]["chart_audits"][0]
        self.assertEqual(audit["distinct_resolved_series_color_count"], 1)
        self.assertTrue(audit["color_review_recommended"])

    def test_governed_peer_series_colors_are_distinct(self) -> None:
        workbook, result = self._workbook_with_series_colors(["24445D", "6F96B8"])

        receipt = validate_style.validate_workbook(workbook, [result])

        self.assertEqual(receipt["status"], "pass")
        audit = receipt["checks"]["chart_audits"][0]
        self.assertEqual(audit["series_colors"], ["#24445D", "#6F96B8"])
        self.assertEqual(audit["distinct_resolved_series_color_count"], 2)
        self.assertFalse(audit["color_review_recommended"])

    def test_wrapped_presentation_context_is_blocked(self) -> None:
        path = self._workbook(show_base=True)
        workbook = load_workbook(path)
        sheet = workbook["分析展示"]
        sheet["A5"] = "统计窗口"
        sheet["B5"] = "2026-07-09 至 2026-07-20"
        sheet["B5"].alignment = Alignment(wrap_text=True)
        workbook.save(path)
        workbook.close()

        receipt = validate_style.validate_workbook(path, [self.result])

        self.assertEqual(receipt["status"], "blocked")
        self.assertEqual(receipt["blockers"][0]["rule_id"], "VIS-LAYOUT-002")
        self.assertIn("B5", receipt["blockers"][0]["message"])

    def test_single_line_presentation_context_passes(self) -> None:
        path = self._workbook(show_base=True)
        workbook = load_workbook(path)
        sheet = workbook["分析展示"]
        sheet["A5"] = "统计窗口"
        sheet["B5"] = "2026-07-09 至 2026-07-20"
        sheet["B5"].alignment = Alignment(wrap_text=False)
        workbook.save(path)
        workbook.close()

        receipt = validate_style.validate_workbook(path, [self.result])

        self.assertEqual(receipt["status"], "pass")
        self.assertEqual(receipt["checks"]["single_line_context"]["context_rows"][0]["wrapped_cells"], [])

    def test_deliberately_wrapped_evidence_is_reviewed_not_blindly_blocked(self) -> None:
        path = self._workbook(show_base=True)
        workbook = load_workbook(path)
        sheet = workbook["分析展示"]
        sheet["A5"] = "质量说明"
        sheet["B5"] = "该指标仅覆盖满足完整观察窗口的玩家。"
        sheet["B5"].alignment = Alignment(wrap_text=True)
        workbook.save(path)
        workbook.close()

        receipt = validate_style.validate_workbook(path, [self.result])

        self.assertEqual(receipt["status"], "pass")
        layout = receipt["checks"]["single_line_context"]
        self.assertTrue(layout["review_recommended"])
        self.assertEqual(layout["observations"][0]["context_kind"], "evidence")

    def test_base_after_percentage_is_flagged_for_density_review(self) -> None:
        path = self.root / "base-after-ratio.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "分析展示"
        sheet.append(["分组", "占比", "玩家数"])
        sheet.append(["A", 0.6, 600])
        sheet.append(["B", 0.4, 400])
        chart = BarChart()
        chart.title = "玩家占比 × 分组"
        chart.add_data(Reference(sheet, min_col=2, min_row=1, max_row=3), titles_from_data=True)
        chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=3))
        sheet.add_chart(chart, "E2")
        workbook.save(path)

        receipt = validate_style.validate_workbook(path, [self.result])

        self.assertEqual(receipt["status"], "pass")
        layout = receipt["checks"]["base_coverage"]["display_layout"]
        self.assertTrue(layout["review_recommended"])
        self.assertIn("base_after_ratio", [item["reason"] for item in layout["observations"]])

    def test_absolute_metric_beside_base_and_ratio_is_flagged_not_blindly_blocked(self) -> None:
        result = self.root / "redundant.csv"
        result.write_text(
            "分组,总玩家数,点击人数,点击率\nA,1000,600,0.6\nB,800,320,0.4\n",
            encoding="utf-8-sig",
        )
        path = self.root / "redundant.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "分析展示"
        sheet.append(["分组", "总玩家数", "点击人数", "点击率"])
        sheet.append(["A", 1000, 600, 0.6])
        sheet.append(["B", 800, 320, 0.4])
        chart = BarChart()
        chart.title = "点击率 × 分组"
        chart.add_data(Reference(sheet, min_col=4, min_row=1, max_row=3), titles_from_data=True)
        chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=3))
        sheet.add_chart(chart, "F2")
        workbook.save(path)

        receipt = validate_style.validate_workbook(path, [result])

        self.assertEqual(receipt["status"], "pass")
        base = receipt["checks"]["base_coverage"]
        self.assertEqual(base["available_fields"], ["总玩家数"])
        self.assertIn(
            "redundant_absolute_with_ratio",
            [item["reason"] for item in base["display_layout"]["observations"]],
        )


if __name__ == "__main__":
    unittest.main()
